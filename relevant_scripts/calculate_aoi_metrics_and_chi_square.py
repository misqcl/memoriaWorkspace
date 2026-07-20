from pathlib import Path
import math
import re
import warnings

import numpy as np
import pandas as pd
from scipy.stats import chi2_contingency
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PATHS
# ==========================================================

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

AOI_FILE = BASE_DIR / "semantic_aois_EOCR" / "semantic_aois_verified.csv"
RESULTS_DIR = BASE_DIR / "results"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"
QA_FILE = BASE_DIR / "fixation_summary_QA.xlsx"
QA_SHEET = "Usable"

TODOS_DIR = BASE_DIR / "Todos"
SUMMARY_FILE = TODOS_DIR / "Summary.xlsx"

OUTPUT_DIR = BASE_DIR / "aoi_analysis_EOCR"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LONG_CSV = OUTPUT_DIR / "aoi_metrics_long.csv"
WIDE_CSV = OUTPUT_DIR / "aoi_metrics_wide.csv"
ASSIGNMENTS_CSV = OUTPUT_DIR / "fixation_aoi_assignments.csv"
CHI_CSV = OUTPUT_DIR / "chi_square_cramers_v.csv"
CONTINGENCY_CSV = OUTPUT_DIR / "chi_square_contingency_reported_emotion_x_face_presence.csv"
XLSX_FILE = OUTPUT_DIR / "aoi_analysis_EOCR.xlsx"


# ==========================================================
# CONFIG
# ==========================================================

N_IMAGES = 60
ALPHA = 0.05

VALID_AOI_STATUSES = {"accepted", "manual"}

ALLOW_BLANK_VERIFICATION_STATUS = True

AOI_TYPES = ("face", "text")


# ==========================================================
# HELPERS
# ==========================================================

def clean_name(name):
    return re.sub(r"\s+", "", str(name).strip())


def normalize_emotion(value):
    if pd.isna(value):
        return np.nan
    s = str(value).strip()
    sl = s.lower()

    mapping = {
        "positive": "Positive", "positiva": "Positive", "positivo": "Positive",
        "neutral": "Neutral", "neutra": "Neutral", "neutro": "Neutral",
        "negative": "Negative", "negativa": "Negative", "negativo": "Negative",
        "no responde": "Neutral",
    }
    return mapping.get(sl, s)


def first_existing_column(df, candidates, required=True):
    lower = {str(c).lower(): c for c in df.columns}
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        if candidate.lower() in lower:
            return lower[candidate.lower()]
    if required:
        raise KeyError(
            f"None of these columns were found: {candidates}. "
            f"Available columns: {list(df.columns)}"
        )
    return None


def cramers_v_corrected(table):
    """Bias-corrected Cramer's V (Bergsma/Wicher correction)."""
    observed = np.asarray(table, dtype=float)
    n = observed.sum()
    if n <= 1:
        return np.nan

    chi2, _, _, _ = chi2_contingency(observed, correction=False)
    phi2 = chi2 / n
    r, k = observed.shape

    phi2corr = max(0.0, phi2 - ((k - 1) * (r - 1)) / (n - 1))
    rcorr = r - ((r - 1) ** 2) / (n - 1)
    kcorr = k - ((k - 1) ** 2) / (n - 1)

    denom = min(kcorr - 1, rcorr - 1)
    if denom <= 0:
        return np.nan
    return math.sqrt(phi2corr / denom)


def inside_any_box(x, y, boxes):
    if boxes.empty:
        return False
    return bool(
        (
            (x >= boxes["x1"]) &
            (x <= boxes["x2"]) &
            (y >= boxes["y1"]) &
            (y <= boxes["y2"])
        ).any()
    )


def union_area_rectangles(rectangles):
    rects = []
    for x1, y1, x2, y2 in rectangles:
        x1, x2 = sorted((float(x1), float(x2)))
        y1, y2 = sorted((float(y1), float(y2)))
        if x2 > x1 and y2 > y1:
            rects.append((x1, y1, x2, y2))

    if not rects:
        return 0.0

    xs = sorted(set([r[0] for r in rects] + [r[2] for r in rects]))
    area = 0.0

    for xa, xb in zip(xs[:-1], xs[1:]):
        if xb <= xa:
            continue
        intervals = []
        for x1, y1, x2, y2 in rects:
            if x1 < xb and x2 > xa:
                intervals.append((y1, y2))
        if not intervals:
            continue

        intervals.sort()
        merged = []
        for start, end in intervals:
            if not merged or start > merged[-1][1]:
                merged.append([start, end])
            else:
                merged[-1][1] = max(merged[-1][1], end)

        y_total = sum(end - start for start, end in merged)
        area += (xb - xa) * y_total

    return area


def load_verified_aois():
    if not AOI_FILE.exists():
        raise FileNotFoundError(f"AOI file not found: {AOI_FILE}")

    aoi = pd.read_csv(AOI_FILE)

    required = {
        "trial_index", "filename", "aoi_type",
        "x1", "y1", "x2", "y2", "img_w", "img_h"
    }
    missing = required - set(aoi.columns)
    if missing:
        raise ValueError(f"AOI file missing columns: {sorted(missing)}")

    if "verification_status" in aoi.columns:
        status = aoi["verification_status"].fillna("").astype(str).str.strip().str.lower()
        keep = status.isin(VALID_AOI_STATUSES)
        if ALLOW_BLANK_VERIFICATION_STATUS:
            keep = keep | status.eq("")
        aoi = aoi.loc[keep].copy()

    aoi["aoi_type"] = aoi["aoi_type"].astype(str).str.strip().str.lower()
    aoi = aoi[aoi["aoi_type"].isin(AOI_TYPES)].copy()

    for c in ["x1", "y1", "x2", "y2", "img_w", "img_h"]:
        aoi[c] = pd.to_numeric(aoi[c], errors="coerce")

    aoi = aoi.dropna(subset=["x1", "y1", "x2", "y2", "img_w", "img_h"])
    return aoi


def build_image_metadata(aoi):
    meta_cols = ["trial_index", "filename", "img_w", "img_h"]
    if "original_emotion" in aoi.columns:
        meta_cols.append("original_emotion")

    meta = aoi[meta_cols].drop_duplicates("filename").copy()

    if IMAGE_AOI_FILE.exists():
        img = pd.read_csv(IMAGE_AOI_FILE)
        rename = {}
        if "emotion" in img.columns and "original_emotion" not in img.columns:
            rename["emotion"] = "original_emotion"
        img = img.rename(columns=rename)

        keep = [c for c in ["trial_index", "filename", "original_emotion"] if c in img.columns]
        img = img[keep].drop_duplicates("filename")

        meta = img.merge(meta, on=[c for c in ["trial_index", "filename", "original_emotion"]
                                  if c in img.columns and c in meta.columns],
                         how="left")

        dims = aoi[["filename", "img_w", "img_h"]].drop_duplicates("filename")
        meta = meta.drop(columns=[c for c in ["img_w", "img_h"] if c in meta.columns], errors="ignore")
        meta = meta.merge(dims, on="filename", how="left")

    if "original_emotion" in meta.columns:
        meta["original_emotion"] = meta["original_emotion"].apply(normalize_emotion)

    return meta.sort_values("trial_index").reset_index(drop=True)


def build_aoi_type_summary(aoi, image_meta):
    rows = []
    for _, img in image_meta.iterrows():
        filename = img["filename"]
        img_w = float(img["img_w"]) if pd.notna(img.get("img_w")) else np.nan
        img_h = float(img["img_h"]) if pd.notna(img.get("img_h")) else np.nan
        image_area = img_w * img_h if pd.notna(img_w) and pd.notna(img_h) else np.nan

        for aoi_type in AOI_TYPES:
            boxes = aoi[(aoi["filename"] == filename) & (aoi["aoi_type"] == aoi_type)]
            rects = boxes[["x1", "y1", "x2", "y2"]].to_numpy().tolist()
            union_area = union_area_rectangles(rects)
            area_prop = union_area / image_area if image_area and image_area > 0 else np.nan

            row = {
                "trial_index": int(img["trial_index"]),
                "filename": filename,
                "aoi_type": aoi_type,
                "aoi_present": int(len(boxes) > 0),
                "aoi_count": int(len(boxes)),
                "aoi_union_area_px": union_area,
                "aoi_union_area_prop": area_prop,
            }
            if "original_emotion" in image_meta.columns:
                row["original_emotion"] = img.get("original_emotion", np.nan)
            rows.append(row)

    return pd.DataFrame(rows)


def load_usable_matrix():
    if not QA_FILE.exists():
        warnings.warn(f"QA file not found: {QA_FILE}; all available participant-trials will be used.")
        return None, None

    usable = pd.read_excel(QA_FILE, sheet_name=QA_SHEET, usecols="B:AM", nrows=N_IMAGES)
    participants = list(usable.columns)
    return usable, participants


def load_response_emotions(participants):
    if not SUMMARY_FILE.exists():
        warnings.warn(f"Summary file not found: {SUMMARY_FILE}; reported emotion cannot be joined.")
        return pd.DataFrame(columns=["participant", "filename", "reported_emotion"])

    users = pd.read_excel(SUMMARY_FILE, header=None, usecols="A", skiprows=1, nrows=len(participants))
    users.columns = ["raw_name"]
    users["participant"] = [f"U{i}" for i in range(1, len(users) + 1)]
    users["clean_name"] = users["raw_name"].apply(clean_name)

    rows = []
    for _, user in users.iterrows():
        participant = user["participant"]
        path = TODOS_DIR / f"respuestas_emociones_{user['clean_name']}.csv"
        if not path.exists():
            continue

        df = pd.read_csv(path)
        filename_col = first_existing_column(df, ["Archivo", "filename"], required=False)
        emotion_col = first_existing_column(df, ["Emocion", "response_emotion"], required=False)
        if filename_col is None or emotion_col is None:
            continue

        tmp = df[[filename_col, emotion_col]].copy()
        tmp.columns = ["filename", "reported_emotion"]
        tmp["participant"] = participant
        tmp["reported_emotion"] = tmp["reported_emotion"].apply(normalize_emotion)
        rows.append(tmp)

    if not rows:
        return pd.DataFrame(columns=["participant", "filename", "reported_emotion"])

    return pd.concat(rows, ignore_index=True).drop_duplicates(["participant", "filename"])


def load_surface_bounds():
    if not IMAGE_AOI_FILE.exists():
        raise FileNotFoundError(f"Image surface AOI file not found: {IMAGE_AOI_FILE}")
    df = pd.read_csv(IMAGE_AOI_FILE)
    needed = {"filename", "surf_left", "surf_right", "surf_top", "surf_bottom"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"Image AOI file missing surface bounds: {sorted(missing)}")
    return df.drop_duplicates("filename").set_index("filename")[
        ["surf_left", "surf_right", "surf_top", "surf_bottom"]
    ].to_dict("index")


def fixation_to_image_pixels(fix, img_w, img_h, bounds):
    """Map Surface Tracker normalized coordinates into reference-image pixel coordinates."""
    sx = float(fix["norm_pos_x"])
    sy = float(fix["norm_pos_y"])

    left = float(bounds["surf_left"])
    right = float(bounds["surf_right"])
    top = float(bounds["surf_top"])
    bottom = float(bounds["surf_bottom"])

    if right <= left or bottom <= top:
        return np.nan, np.nan

    ix = (sx - left) / (right - left)
    iy = (sy - top) / (bottom - top)

    if ix < 0 or ix > 1 or iy < 0 or iy > 1:
        return np.nan, np.nan

    return ix * (img_w - 1), iy * (img_h - 1)


def main():
    print("Loading verified semantic AOIs...")
    aoi = load_verified_aois()
    image_meta = build_image_metadata(aoi)
    aoi_summary = build_aoi_type_summary(aoi, image_meta)

    usable, qa_participants = load_usable_matrix()
    if qa_participants is None:
        participants = sorted([p.name for p in RESULTS_DIR.iterdir() if p.is_dir()])
    else:
        participants = qa_participants

    response_emotions = load_response_emotions(participants)
    surface_bounds = load_surface_bounds()

    metric_rows = []
    assignment_rows = []

    print(f"Processing {len(participants)} participants...")

    for p_idx, participant in enumerate(participants, start=1):
        participant_dir = RESULTS_DIR / str(participant)
        if not participant_dir.exists():
            print(f"{p_idx:02d}/{len(participants)} {participant}: missing folder")
            continue

        candidates = sorted(participant_dir.glob("*fixation*.csv"))
        if not candidates:
            print(f"{p_idx:02d}/{len(participants)} {participant}: no fixation CSV")
            continue

        fix = pd.read_csv(candidates[0])

        required_fix = {"filename", "fixation_id", "start_timestamp", "duration", "norm_pos_x", "norm_pos_y"}
        missing = required_fix - set(fix.columns)
        if missing:
            raise ValueError(f"{candidates[0]} missing columns: {sorted(missing)}")

        fix["duration"] = pd.to_numeric(fix["duration"], errors="coerce")
        fix["start_timestamp"] = pd.to_numeric(fix["start_timestamp"], errors="coerce")
        fix["norm_pos_x"] = pd.to_numeric(fix["norm_pos_x"], errors="coerce")
        fix["norm_pos_y"] = pd.to_numeric(fix["norm_pos_y"], errors="coerce")
        fix = fix.dropna(subset=["duration", "start_timestamp", "norm_pos_x", "norm_pos_y"])

        participant_rows = 0

        for _, img in image_meta.iterrows():
            trial = int(img["trial_index"])
            filename = img["filename"]

            if usable is not None:
                try:
                    if usable.loc[trial - 1, participant] != 1:
                        continue
                except Exception:
                    continue

            img_fix = fix[fix["filename"].astype(str) == str(filename)].copy()
            if img_fix.empty:
                continue

            img_w = float(img["img_w"])
            img_h = float(img["img_h"])

            if filename not in surface_bounds:
                warnings.warn(f"No surface bounds for {filename}; skipping.")
                continue
            bounds = surface_bounds[filename]

            trial_start = img_fix["start_timestamp"].min()

            total_fixations = len(img_fix)
            total_dwell_ms = img_fix["duration"].sum()

            for aoi_type in AOI_TYPES:
                boxes = aoi[
                    (aoi["filename"].astype(str) == str(filename)) &
                    (aoi["aoi_type"] == aoi_type)
                ]

                hits = []
                hit_start_times = []
                hit_durations = []

                for _, f in img_fix.iterrows():
                    x_px, y_px = fixation_to_image_pixels(f, img_w, img_h, bounds)
                    hit = False if pd.isna(x_px) or pd.isna(y_px) else inside_any_box(x_px, y_px, boxes)
                    hits.append(hit)

                    assignment_rows.append({
                        "participant": participant,
                        "trial_index": trial,
                        "filename": filename,
                        "fixation_id": f["fixation_id"],
                        "aoi_type": aoi_type,
                        "inside_aoi": int(hit),
                        "x_px": x_px,
                        "y_px": y_px,
                        "start_timestamp": f["start_timestamp"],
                        "duration_ms": f["duration"],
                    })

                    if hit:
                        hit_start_times.append(float(f["start_timestamp"]))
                        hit_durations.append(float(f["duration"]))

                aoi_fixation_count = int(sum(hits))
                aoi_dwell_ms = float(sum(hit_durations))

                fixation_proportion = (
                    aoi_fixation_count / total_fixations if total_fixations > 0 else np.nan
                )
                dwell_proportion = (
                    aoi_dwell_ms / total_dwell_ms if total_dwell_ms > 0 else np.nan
                )

                ttff_ms = (
                    (min(hit_start_times) - trial_start) * 1000.0
                    if hit_start_times else np.nan
                )

                summary_row = aoi_summary[
                    (aoi_summary["filename"].astype(str) == str(filename)) &
                    (aoi_summary["aoi_type"] == aoi_type)
                ].iloc[0]

                area_prop = float(summary_row["aoi_union_area_prop"])
                area_normalized_density = (
                    fixation_proportion / area_prop
                    if area_prop > 0 else np.nan
                )

                row = {
                    "participant": participant,
                    "trial_index": trial,
                    "filename": filename,
                    "aoi_type": aoi_type,
                    "aoi_present": int(summary_row["aoi_present"]),
                    "aoi_count": int(summary_row["aoi_count"]),
                    "aoi_union_area_px": summary_row["aoi_union_area_px"],
                    "aoi_union_area_prop": area_prop,
                    "total_trial_fixations": total_fixations,
                    "total_trial_dwell_ms": total_dwell_ms,
                    "aoi_fixation_count": aoi_fixation_count,
                    "aoi_dwell_ms": aoi_dwell_ms,
                    "fixation_proportion": fixation_proportion,
                    "dwell_proportion": dwell_proportion,
                    "ttff_ms": ttff_ms,
                    "area_normalized_density": area_normalized_density,
                }

                if "original_emotion" in image_meta.columns:
                    row["original_emotion"] = img.get("original_emotion", np.nan)

                metric_rows.append(row)
                participant_rows += 1

        print(f"{p_idx:02d}/{len(participants)} {participant}: {participant_rows} metric rows")

    metrics = pd.DataFrame(metric_rows)
    assignments = pd.DataFrame(assignment_rows)

    if metrics.empty:
        raise RuntimeError("No AOI metric rows were produced. Check paths and fixation filenames.")

    metrics = metrics.merge(
        response_emotions,
        on=["participant", "filename"],
        how="left"
    )

    metrics.to_csv(LONG_CSV, index=False, encoding="utf-8-sig")
    assignments.to_csv(ASSIGNMENTS_CSV, index=False, encoding="utf-8-sig")

    value_cols = [
        "fixation_proportion",
        "dwell_proportion",
        "ttff_ms",
        "area_normalized_density",
        "aoi_fixation_count",
        "aoi_dwell_ms",
    ]

    id_cols = [
        c for c in
        ["participant", "trial_index", "filename", "original_emotion", "reported_emotion"]
        if c in metrics.columns
    ]

    wide = metrics.pivot_table(
        index=id_cols,
        columns="aoi_type",
        values=value_cols,
        aggfunc="first"
    )
    wide.columns = [f"{aoi_type}_{metric}" for metric, aoi_type in wide.columns]
    wide = wide.reset_index()
    wide.to_csv(WIDE_CSV, index=False, encoding="utf-8-sig")

    face_metrics = metrics[metrics["aoi_type"] == "face"].copy()
    chi_rows = []
    contingency = pd.DataFrame()

    if "reported_emotion" in face_metrics.columns:
        chi_data = face_metrics.dropna(subset=["reported_emotion", "aoi_present"]).copy()

        if not chi_data.empty:
            contingency = pd.crosstab(
                chi_data["reported_emotion"],
                chi_data["aoi_present"]
            )
            contingency = contingency.rename(columns={0: "Face absent", 1: "Face present"})

            if contingency.shape[0] >= 2 and contingency.shape[1] >= 2:
                chi2, p, dof, expected = chi2_contingency(contingency, correction=False)
                v = cramers_v_corrected(contingency.to_numpy())

                expected_arr = np.asarray(expected)
                chi_rows.append({
                    "association": "reported_emotion_x_face_presence",
                    "chi_square": chi2,
                    "degrees_of_freedom": dof,
                    "p_value": p,
                    "cramers_v_bias_corrected": v,
                    "n": int(contingency.to_numpy().sum()),
                    "min_expected_count": float(expected_arr.min()),
                    "cells_expected_lt_5": int((expected_arr < 5).sum()),
                    "significant_0_05": bool(p < ALPHA),
                })

    chi_results = pd.DataFrame(chi_rows)
    chi_results.to_csv(CHI_CSV, index=False, encoding="utf-8-sig")
    contingency.to_csv(CONTINGENCY_CSV, encoding="utf-8-sig")

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="AOI Metrics Long", index=False)
        wide.to_excel(writer, sheet_name="AOI Metrics Wide", index=False)
        aoi_summary.to_excel(writer, sheet_name="AOI Area Summary", index=False)
        chi_results.to_excel(writer, sheet_name="Chi-square Cramers V", index=False)
        contingency.to_excel(writer, sheet_name="Contingency Face")
        response_emotions.to_excel(writer, sheet_name="Response Emotions", index=False)

    style_excel(XLSX_FILE)

    print("\nDone.")
    print(f"Long metrics:       {LONG_CSV}")
    print(f"Wide metrics:       {WIDE_CSV}")
    print(f"Assignments:        {ASSIGNMENTS_CSV}")
    print(f"Chi-square results: {CHI_CSV}")
    print(f"Excel workbook:     {XLSX_FILE}")


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    wb.save(path)


if __name__ == "__main__":
    main()
