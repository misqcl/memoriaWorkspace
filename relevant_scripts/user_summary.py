import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

RESULTS_FOLDER = BASE_DIR / "results"
filename = "fixation_summary.xlsx"
OUTPUT_FILE = os.path.join(BASE_DIR,filename)
wb = Workbook()

# ======================================================
# Participants sheet
# ======================================================

ws = wb.active
ws.title = "Participants"

headers = [
    "Participant",
    "Total Fixations",
    "Total Fixation Duration (s)",
    "Mean Fixation Duration (ms)",
]

headers.extend(
    [f"Image {i}" for i in range(1, 61)]
)

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

summary = []

row = 2

for participant in sorted(os.listdir(RESULTS_FOLDER)):

    participant_folder = os.path.join(
        RESULTS_FOLDER,
        participant
    )

    if not os.path.isdir(participant_folder):
        continue

    csv_file = os.path.join(
        participant_folder,
        "valid_fixations.csv"
    )

    if not os.path.exists(csv_file):
        continue

    df = pd.read_csv(csv_file)

    total_fix = len(df)
    total_duration = df["duration"].sum() / 1000
    mean_duration = df["duration"].mean()

    fix_per_image = (
        df.groupby("trial")
        .size()
        .reindex(range(1, 61), fill_value=0)
    )

    ws.cell(row=row, column=1, value=participant)
    ws.cell(row=row, column=2, value=total_fix)
    ws.cell(row=row, column=3, value=total_duration)
    ws.cell(row=row, column=4, value=mean_duration)

    for i, value in enumerate(fix_per_image, start=5):  
        ws.cell(row=row, column=i, value=int(value))

    summary.append({
        "Participant": participant,
        "Fixations": total_fix,
        "Duration": total_duration
    })

    row += 1

# ======================================================
# Summary sheet
# ======================================================

summary_df = pd.DataFrame(summary)

summary_ws = wb.create_sheet("Summary")

summary_ws["A1"] = "Metric"
summary_ws["B1"] = "Value"

summary_ws["A1"].font = Font(bold=True)
summary_ws["B1"].font = Font(bold=True)

least_fix = summary_df.loc[
    summary_df["Fixations"].idxmin()
]

most_fix = summary_df.loc[
    summary_df["Fixations"].idxmax()
]

most_duration = summary_df.loc[
    summary_df["Duration"].idxmax()
]

metrics = [

    ("Participants", len(summary_df)),

    ("Average Fixations",
     round(summary_df["Fixations"].mean(), 2)),

    ("Average Fixation Duration (s)",
     round(summary_df["Duration"].mean(), 2)),

    ("Fewest Fixations",
     f"{least_fix['Participant']} ({least_fix['Fixations']})"),

    ("Most Fixations",
     f"{most_fix['Participant']} ({most_fix['Fixations']})"),

    ("Longest Total Fixation Duration",
     f"{most_duration['Participant']} ({most_duration['Duration']:.2f} s)")
]

for i, (metric, value) in enumerate(metrics, start=2):

    summary_ws.cell(row=i, column=1, value=metric)
    summary_ws.cell(row=i, column=2, value=value)

# ======================================================
# Auto-size columns
# ======================================================

for sheet in wb.worksheets:

    for column in sheet.columns:

        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column
        )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = length + 2

# ======================================================

wb.save(OUTPUT_FILE)

print(f"Saved {OUTPUT_FILE}")