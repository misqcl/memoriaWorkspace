from pathlib import Path
import shutil
import warnings

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PATHS
# ==========================================================

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

IMAGE_ROOT = BASE_DIR / "imagenes" / "top_imagenes"
AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"


OUTPUT_DIR = BASE_DIR / "semantic_aois_EOCR"
PREVIEW_DIR = OUTPUT_DIR / "previews"

AOI_OUTPUT_CSV = OUTPUT_DIR / "semantic_aois_raw.csv"
AOI_VERIFIED_CSV = OUTPUT_DIR / "semantic_aois_verified.csv"
SUMMARY_XLSX = OUTPUT_DIR / "detection_summary.xlsx"


# ==========================================================
# CONFIG
# ==========================================================

DETECT_FACES = True
DETECT_TEXT = True

# MediaPipe face detector parameters.
MEDIAPIPE_MIN_DETECTION_CONFIDENCE = 0.60
MEDIAPIPE_MODEL_SELECTION = 1  # 0 = short range, 1 = full range
FACE_PADDING_PCT = 0.08
MIN_FACE_AREA_PROP = 0.0025
MAX_FACE_AREA_PROP = 0.45
FACE_MIN_ASPECT_RATIO = 0.55
FACE_MAX_ASPECT_RATIO = 1.80

# EasyOCR parameters.
# Use ["en"] for English ads. Add another language code only if your stimuli require it.
EASYOCR_LANGS = ["en"]
EASYOCR_GPU = False  # Set True only if your torch/CUDA setup supports it.
EASYOCR_MIN_CONFIDENCE = 0.30
TEXT_PADDING_PX = 6

MIN_AOI_AREA_PX = 50
BOX_MERGE_IOU_THRESHOLD = 0.25


# ==========================================================
# OPTIONAL EASYOCR IMPORT
# ==========================================================

try:
    import easyocr
    EASYOCR_AVAILABLE = True
except Exception:
    easyocr = None
    EASYOCR_AVAILABLE = False

_EASYOCR_READER = None

def get_easyocr_reader():
    global _EASYOCR_READER

    if not EASYOCR_AVAILABLE:
        return None

    if _EASYOCR_READER is None:
        _EASYOCR_READER = easyocr.Reader(EASYOCR_LANGS, gpu=EASYOCR_GPU)

    return _EASYOCR_READER


# ==========================================================
# OPTIONAL MEDIAPIPE IMPORT
# ==========================================================

try:
    import mediapipe as mp
    MEDIAPIPE_AVAILABLE = hasattr(mp, "solutions")
except Exception:
    mp = None
    MEDIAPIPE_AVAILABLE = False


# ==========================================================
# HELPERS
# ==========================================================

def ensure_dirs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)


def style_excel(path: Path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    wb.save(path)


def find_image_files(image_root: Path):
    image_paths = {}

    for path in image_root.rglob("*"):
        if path.suffix.lower() in [".jpg", ".jpeg", ".png"]:
            image_paths[path.name] = path

    return image_paths


def clamp_box(x1, y1, x2, y2, w, h):
    x1 = max(0, min(int(round(x1)), w - 1))
    y1 = max(0, min(int(round(y1)), h - 1))
    x2 = max(0, min(int(round(x2)), w))
    y2 = max(0, min(int(round(y2)), h))

    if x2 <= x1:
        x2 = min(w, x1 + 1)

    if y2 <= y1:
        y2 = min(h, y1 + 1)

    return x1, y1, x2, y2


def pad_box(x1, y1, x2, y2, pad_x, pad_y, w, h):
    return clamp_box(
        x1 - pad_x,
        y1 - pad_y,
        x2 + pad_x,
        y2 + pad_y,
        w,
        h,
    )


def box_area(box):
    x1, y1, x2, y2 = box
    return max(0, x2 - x1) * max(0, y2 - y1)


def passes_face_box_filters(box, img_w, img_h):
    x1, y1, x2, y2 = box
    bw = x2 - x1
    bh = y2 - y1

    if bw <= 0 or bh <= 0:
        return False

    area_prop = (bw * bh) / (img_w * img_h)
    aspect = bw / bh

    if area_prop < MIN_FACE_AREA_PROP:
        return False

    if area_prop > MAX_FACE_AREA_PROP:
        return False

    if aspect < FACE_MIN_ASPECT_RATIO or aspect > FACE_MAX_ASPECT_RATIO:
        return False

    return True


def iou(box_a, box_b):
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b

    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)

    inter = box_area((ix1, iy1, ix2, iy2))
    union = box_area(box_a) + box_area(box_b) - inter

    if union <= 0:
        return 0.0

    return inter / union


def merge_boxes(boxes, threshold=0.25):
    if not boxes:
        return []

    boxes = [dict(b) for b in boxes]
    changed = True

    while changed:
        changed = False
        merged = []
        used = [False] * len(boxes)

        for i in range(len(boxes)):
            if used[i]:
                continue

            current = boxes[i]
            current_box = (
                current["x1"],
                current["y1"],
                current["x2"],
                current["y2"],
            )

            for j in range(i + 1, len(boxes)):
                if used[j]:
                    continue

                other = boxes[j]
                other_box = (
                    other["x1"],
                    other["y1"],
                    other["x2"],
                    other["y2"],
                )

                if iou(current_box, other_box) >= threshold:
                    current["x1"] = min(current["x1"], other["x1"])
                    current["y1"] = min(current["y1"], other["y1"])
                    current["x2"] = max(current["x2"], other["x2"])
                    current["y2"] = max(current["y2"], other["y2"])
                    current["confidence"] = max(
                        float(current.get("confidence", 0) or 0),
                        float(other.get("confidence", 0) or 0),
                    )

                    if current.get("text") or other.get("text"):
                        current["text"] = " ".join(
                            str(x).strip()
                            for x in [current.get("text", ""), other.get("text", "")]
                            if str(x).strip()
                        )

                    used[j] = True
                    changed = True

            used[i] = True
            merged.append(current)

        boxes = merged

    return boxes


def box_to_record(
    trial_index,
    original_emotion,
    filename,
    image_path,
    img_w,
    img_h,
    aoi_type,
    box,
    confidence=np.nan,
    source="",
    text="",
):
    x1, y1, x2, y2 = box
    width = x2 - x1
    height = y2 - y1
    area_px = width * height

    return {
        "trial_index": int(trial_index),
        "original_emotion": original_emotion,
        "filename": filename,
        "image_path": str(image_path),
        "img_w": img_w,
        "img_h": img_h,
        "aoi_type": aoi_type,
        "x1": x1,
        "y1": y1,
        "x2": x2,
        "y2": y2,
        "width": width,
        "height": height,
        "area_px": area_px,
        "area_prop": area_px / (img_w * img_h) if img_w * img_h > 0 else np.nan,
        "confidence": confidence,
        "source": source,
        "text": text,
        "verification_status": "needs_review",
        "verification_notes": "",
    }


# ==========================================================
# FACE DETECTION: MEDIAPIPE
# ==========================================================

def detect_face_aois(image_bgr):
    if not DETECT_FACES:
        return [], "disabled"

    if not MEDIAPIPE_AVAILABLE:
        return [], "skipped_no_mediapipe"

    h, w = image_bgr.shape[:2]
    image_rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    boxes = []

    with mp.solutions.face_detection.FaceDetection(
        model_selection=MEDIAPIPE_MODEL_SELECTION,
        min_detection_confidence=MEDIAPIPE_MIN_DETECTION_CONFIDENCE,
    ) as detector:
        results = detector.process(image_rgb)

    if not results.detections:
        return [], "mediapipe_face_detection"

    for detection in results.detections:
        score = float(detection.score[0]) if detection.score else np.nan
        rel_box = detection.location_data.relative_bounding_box

        x1 = rel_box.xmin * w
        y1 = rel_box.ymin * h
        x2 = (rel_box.xmin + rel_box.width) * w
        y2 = (rel_box.ymin + rel_box.height) * h

        bw = x2 - x1
        bh = y2 - y1

        x1, y1, x2, y2 = pad_box(
            x1,
            y1,
            x2,
            y2,
            bw * FACE_PADDING_PCT,
            bh * FACE_PADDING_PCT,
            w,
            h,
        )

        box = (x1, y1, x2, y2)

        if not passes_face_box_filters(box, w, h):
            continue

        if box_area(box) < MIN_AOI_AREA_PX:
            continue

        boxes.append({
            "x1": x1,
            "y1": y1,
            "x2": x2,
            "y2": y2,
            "confidence": score,
            "source": "mediapipe_face_detection",
            "text": "",
        })

    return merge_boxes(boxes, BOX_MERGE_IOU_THRESHOLD), "mediapipe_face_detection"

# ==========================================================
# TEXT DETECTION / OCR: EASYOCR
# ==========================================================

def detect_text_aois(image_bgr):
    if not EASYOCR_AVAILABLE:
        return []

    reader = get_easyocr_reader()
    if reader is None:
        return []

    h, w = image_bgr.shape[:2]
    image_rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)

    try:
        # EasyOCR returns: [([[x,y],...4 points], text, confidence), ...]
        results = reader.readtext(image_rgb, detail=1, paragraph=False)
    except Exception as exc:
        warnings.warn(f"EasyOCR failed: {exc}")
        return []

    boxes = []

    for item in results:
        if len(item) < 3:
            continue

        points, detected_text, confidence = item
        confidence = float(confidence)

        if confidence < EASYOCR_MIN_CONFIDENCE:
            continue

        detected_text = str(detected_text).strip()
        if not detected_text:
            continue

        pts = np.array(points, dtype=float)
        x1 = float(np.min(pts[:, 0]))
        y1 = float(np.min(pts[:, 1]))
        x2 = float(np.max(pts[:, 0]))
        y2 = float(np.max(pts[:, 1]))

        x1, y1, x2, y2 = pad_box(
            x1,
            y1,
            x2,
            y2,
            TEXT_PADDING_PX,
            TEXT_PADDING_PX,
            w,
            h,
        )

        if box_area((x1, y1, x2, y2)) < MIN_AOI_AREA_PX:
            continue

        boxes.append({
            "x1": x1,
            "y1": y1,
            "x2": x2,
            "y2": y2,
            "confidence": confidence,
            "source": "easyocr",
            "text": detected_text,
        })

    return merge_boxes(boxes, BOX_MERGE_IOU_THRESHOLD)


# ==========================================================
# PREVIEW EXPORT
# ==========================================================

def draw_previews(image_bgr, records, output_path):
    preview = image_bgr.copy()

    for rec in records:
        x1 = int(rec["x1"])
        y1 = int(rec["y1"])
        x2 = int(rec["x2"])
        y2 = int(rec["y2"])

        if rec["aoi_type"] == "face":
            color = (0, 255, 0)
        elif rec["aoi_type"] == "text":
            color = (255, 0, 0)
        else:
            color = (0, 255, 255)

        cv2.rectangle(preview, (x1, y1), (x2, y2), color, 2)

        label = rec["aoi_type"]

        if rec["aoi_type"] == "text" and str(rec.get("text", "")).strip():
            label = f"text: {str(rec['text'])[:20]}"

        if rec["aoi_type"] == "face":
            conf = rec.get("confidence", np.nan)
            if pd.notna(conf):
                label = f"face {conf:.2f}"

        cv2.putText(
            preview,
            label,
            (x1, max(15, y1 - 5)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            color,
            1,
            cv2.LINE_AA,
        )

    cv2.imwrite(str(output_path), preview)


# ==========================================================
# MAIN
# ==========================================================

def main():
    ensure_dirs()

    if not IMAGE_ROOT.exists():
        raise FileNotFoundError(f"Image folder not found: {IMAGE_ROOT}")

    if not AOI_FILE.exists():
        raise FileNotFoundError(f"AOI file not found: {AOI_FILE}")

    if DETECT_FACES and not MEDIAPIPE_AVAILABLE:
        print("WARNING: MediaPipe is not installed or does not expose mp.solutions.")
        print("Face AOI detection will be skipped.")
        print("Install/use a compatible Python environment, e.g. Python 3.12 with mediapipe.")
        print("Text AOI detection will still run if EasyOCR is available.")

    if DETECT_TEXT and not EASYOCR_AVAILABLE:
        print("WARNING: EasyOCR is not installed or could not be imported.")
        print("Text AOI detection will be skipped.")
        print("Install with: python -m pip install easyocr torch torchvision")

    aoi = pd.read_csv(AOI_FILE)

    if "emotion" in aoi.columns:
        aoi = aoi.rename(columns={"emotion": "original_emotion"})

    image_lookup = find_image_files(IMAGE_ROOT)

    all_records = []
    summary_rows = []

    for _, row in aoi.sort_values("trial_index").iterrows():
        filename = row["filename"]

        if filename not in image_lookup:
            print(f"Missing image: {filename}")
            summary_rows.append({
                "trial_index": row["trial_index"],
                "filename": filename,
                "status": "missing_image",
                "face_detector_used": "",
                "face_aois": 0,
                "text_aois": 0,
                "total_aois": 0,
            })
            continue

        image_path = image_lookup[filename]
        image_bgr = cv2.imread(str(image_path))

        if image_bgr is None:
            print(f"Could not read image: {image_path}")
            summary_rows.append({
                "trial_index": row["trial_index"],
                "filename": filename,
                "status": "read_error",
                "face_detector_used": "",
                "face_aois": 0,
                "text_aois": 0,
                "total_aois": 0,
            })
            continue

        img_h, img_w = image_bgr.shape[:2]

        image_records = []

        face_boxes, face_detector_used = detect_face_aois(image_bgr)

        for box in face_boxes:
            rec = box_to_record(
                trial_index=row["trial_index"],
                original_emotion=row["original_emotion"],
                filename=filename,
                image_path=image_path,
                img_w=img_w,
                img_h=img_h,
                aoi_type="face",
                box=(box["x1"], box["y1"], box["x2"], box["y2"]),
                confidence=box.get("confidence", np.nan),
                source=box.get("source", face_detector_used),
                text="",
            )
            image_records.append(rec)

        if DETECT_TEXT:
            text_boxes = detect_text_aois(image_bgr)

            for box in text_boxes:
                rec = box_to_record(
                    trial_index=row["trial_index"],
                    original_emotion=row["original_emotion"],
                    filename=filename,
                    image_path=image_path,
                    img_w=img_w,
                    img_h=img_h,
                    aoi_type="text",
                    box=(box["x1"], box["y1"], box["x2"], box["y2"]),
                    confidence=box.get("confidence", np.nan),
                    source=box.get("source", "easyocr"),
                    text=box.get("text", ""),
                )
                image_records.append(rec)

        # Assign stable AOI IDs per image and AOI type.
        type_counts = {}

        for rec in image_records:
            aoi_type = rec["aoi_type"]
            type_counts[aoi_type] = type_counts.get(aoi_type, 0) + 1
            rec["aoi_id"] = f"{Path(filename).stem}_{aoi_type}_{type_counts[aoi_type]:02d}"

        all_records.extend(image_records)

        preview_path = PREVIEW_DIR / f"{int(row['trial_index']):02d}_{Path(filename).stem}_semantic_aois.png"
        draw_previews(image_bgr, image_records, preview_path)

        face_count = sum(1 for r in image_records if r["aoi_type"] == "face")
        text_count = sum(1 for r in image_records if r["aoi_type"] == "text")

        summary_rows.append({
            "trial_index": int(row["trial_index"]),
            "filename": filename,
            "status": "ok",
            "face_detector_used": face_detector_used,
            "face_aois": face_count,
            "text_aois": text_count,
            "total_aois": len(image_records),
            "preview_path": str(preview_path),
        })

        print(
            f"{int(row['trial_index']):02d}/60 "
            f"{filename}: faces={face_count}, text={text_count}, "
            f"face_detector={face_detector_used}"
        )

    output_columns = [
        "trial_index",
        "original_emotion",
        "filename",
        "aoi_id",
        "aoi_type",
        "x1",
        "y1",
        "x2",
        "y2",
        "width",
        "height",
        "area_px",
        "area_prop",
        "confidence",
        "source",
        "text",
        "verification_status",
        "verification_notes",
        "img_w",
        "img_h",
        "image_path",
    ]

    if all_records:
        aoi_df = pd.DataFrame(all_records)
        aoi_df = aoi_df[output_columns]
    else:
        aoi_df = pd.DataFrame(columns=output_columns)

    summary_df = pd.DataFrame(summary_rows)

    aoi_df.to_csv(AOI_OUTPUT_CSV, index=False, encoding="utf-8-sig")

    # Create verified copy only if it does not already exist.
    # This prevents overwriting manual corrections.
    if not AOI_VERIFIED_CSV.exists():
        shutil.copyfile(AOI_OUTPUT_CSV, AOI_VERIFIED_CSV)
        verified_message = "Created verified AOI file from raw detections."
    else:
        verified_message = "Verified AOI file already exists; manual edits were not overwritten."

    with pd.ExcelWriter(SUMMARY_XLSX, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Detection Summary", index=False)
        aoi_df.to_excel(writer, sheet_name="Semantic AOIs Raw", index=False)

    style_excel(SUMMARY_XLSX)

    print("\nDone.")
    print(f"Raw AOIs:      {AOI_OUTPUT_CSV}")
    print(f"Verified AOIs: {AOI_VERIFIED_CSV}")
    print(f"Previews:      {PREVIEW_DIR}")
    print(f"Summary XLSX:  {SUMMARY_XLSX}")
    print(verified_message)
    print("\nImportant: review preview images and manually correct semantic_aois_verified.csv before AOI statistics.")


if __name__ == "__main__":
    main()
