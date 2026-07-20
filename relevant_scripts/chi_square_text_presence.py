from pathlib import Path
import math
import re
import numpy as np
import pandas as pd
from scipy.stats import chi2_contingency
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PATHS
# ==========================================================

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

AOI_FILE = BASE_DIR / "semantic_aois_EOCR" / "semantic_aois_verified.csv"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"
QA_FILE = BASE_DIR / "fixation_summary_QA.xlsx"
QA_SHEET = "Usable"

TODOS_DIR = BASE_DIR / "Todos"
SUMMARY_FILE = TODOS_DIR / "Summary.xlsx"

OUTPUT_DIR = BASE_DIR / "aoi_analysis_EOCR"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

RESULT_CSV = OUTPUT_DIR / "chi_square_text_presence.csv"
CONTINGENCY_CSV = OUTPUT_DIR / "chi_square_contingency_reported_emotion_x_text_presence.csv"
XLSX_FILE = OUTPUT_DIR / "chi_square_text_presence.xlsx"


# ==========================================================
# CONFIG
# ==========================================================

N_IMAGES = 60
VALID_AOI_STATUSES = {"accepted", "manual", "needs_review"}
ALLOW_BLANK_VERIFICATION_STATUS = True
ALPHA = 0.05


# ==========================================================
# HELPERS
# ==========================================================

def clean_name(name):
    return re.sub(r"\s+", "", str(name).strip())


def normalize_emotion(value):
    if pd.isna(value):
        return np.nan

    s = str(value).strip()
    sl = s.lower()

    mapping = {
        "positive": "Positive", "positiva": "Positive", "positivo": "Positive",
        "neutral": "Neutral", "neutra": "Neutral", "neutro": "Neutral",
        "negative": "Negative", "negativa": "Negative", "negativo": "Negative",
        "no responde": "Neutral", "no response": "Neutral", "no_response": "Neutral",
    }

    return mapping.get(sl, s)


def first_existing_column(df, candidates, required=True):
    lower = {str(c).lower(): c for c in df.columns}
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        if candidate.lower() in lower:
            return lower[candidate.lower()]

    if required:
        raise KeyError(
            f"None of these columns were found: {candidates}. "
            f"Available columns: {list(df.columns)}"
        )

    return None


def cramers_v_corrected(table):
    """Bias-corrected Cramer's V."""
    observed = np.asarray(table, dtype=float)
    n = observed.sum()

    if n <= 1:
        return np.nan

    chi2, _, _, _ = chi2_contingency(observed, correction=False)
    phi2 = chi2 / n
    r, k = observed.shape

    phi2corr = max(0.0, phi2 - ((k - 1) * (r - 1)) / (n - 1))
    rcorr = r - ((r - 1) ** 2) / (n - 1)
    kcorr = k - ((k - 1) ** 2) / (n - 1)

    denom = min(kcorr - 1, rcorr - 1)
    if denom <= 0:
        return np.nan

    return math.sqrt(phi2corr / denom)


def load_image_list():
    if not IMAGE_AOI_FILE.exists():
        raise FileNotFoundError(f"Missing image list: {IMAGE_AOI_FILE}")

    df = pd.read_csv(IMAGE_AOI_FILE)
    if "emotion" in df.columns and "original_emotion" not in df.columns:
        df = df.rename(columns={"emotion": "original_emotion"})

    required = {"trial_index", "filename"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Image list missing columns: {sorted(missing)}")

    keep = [c for c in ["trial_index", "filename", "original_emotion"] if c in df.columns]
    out = df[keep].drop_duplicates("filename").copy()

    if "original_emotion" in out.columns:
        out["original_emotion"] = out["original_emotion"].apply(normalize_emotion)

    return out.sort_values("trial_index").reset_index(drop=True)


def load_text_presence_by_image():
    if not AOI_FILE.exists():
        raise FileNotFoundError(f"Missing AOI file: {AOI_FILE}")

    aoi = pd.read_csv(AOI_FILE)

    required = {"filename", "aoi_type"}
    missing = required - set(aoi.columns)
    if missing:
        raise ValueError(f"AOI file missing columns: {sorted(missing)}")

    if "verification_status" in aoi.columns:
        status = aoi["verification_status"].fillna("").astype(str).str.strip().str.lower()
        keep = status.isin(VALID_AOI_STATUSES)
        if ALLOW_BLANK_VERIFICATION_STATUS:
            keep = keep | status.eq("")
        aoi = aoi.loc[keep].copy()

    aoi["aoi_type"] = aoi["aoi_type"].astype(str).str.strip().str.lower()

    text_presence = (
        aoi[aoi["aoi_type"] == "text"]
        .groupby("filename")
        .size()
        .reset_index(name="text_aoi_count")
    )

    text_presence["text_present"] = (text_presence["text_aoi_count"] > 0).astype(int)

    return text_presence


def load_usable_matrix():
    if not QA_FILE.exists():
        raise FileNotFoundError(f"Missing QA file: {QA_FILE}")

    usable = pd.read_excel(QA_FILE, sheet_name=QA_SHEET, usecols="B:AM", nrows=N_IMAGES)
    participants = list(usable.columns)
    return usable, participants


def load_response_emotions(participants):
    if not SUMMARY_FILE.exists():
        raise FileNotFoundError(f"Missing Summary.xlsx: {SUMMARY_FILE}")

    users = pd.read_excel(SUMMARY_FILE, header=None, usecols="A", skiprows=1, nrows=len(participants))
    users.columns = ["raw_name"]
    users["participant"] = [f"U{i}" for i in range(1, len(users) + 1)]
    users["clean_name"] = users["raw_name"].apply(clean_name)

    rows = []

    for _, user in users.iterrows():
        participant = user["participant"]
        path = TODOS_DIR / f"respuestas_emociones_{user['clean_name']}.csv"

        if not path.exists():
            continue

        df = pd.read_csv(path)

        filename_col = first_existing_column(df, ["Archivo", "filename"], required=False)
        emotion_col = first_existing_column(df, ["Emocion", "response_emotion", "reported_emotion"], required=False)

        if filename_col is None or emotion_col is None:
            continue

        tmp = df[[filename_col, emotion_col]].copy()
        tmp.columns = ["filename", "reported_emotion"]
        tmp["participant"] = participant
        tmp["reported_emotion"] = tmp["reported_emotion"].apply(normalize_emotion)

        rows.append(tmp)

    if not rows:
        raise RuntimeError("No participant response-emotion files could be loaded.")

    return pd.concat(rows, ignore_index=True).drop_duplicates(["participant", "filename"])


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    wb.save(path)


# ==========================================================
# MAIN
# ==========================================================

def main():
    images = load_image_list()
    text_presence = load_text_presence_by_image()

    images = images.merge(text_presence, on="filename", how="left")
    images["text_aoi_count"] = images["text_aoi_count"].fillna(0).astype(int)
    images["text_present"] = images["text_present"].fillna(0).astype(int)

    usable, participants = load_usable_matrix()
    responses = load_response_emotions(participants)

    rows = []

    for _, img in images.iterrows():
        trial = int(img["trial_index"])
        filename = img["filename"]

        for participant in participants:
            try:
                if usable.loc[trial - 1, participant] != 1:
                    continue
            except Exception:
                continue

            rows.append({
                "participant": participant,
                "trial_index": trial,
                "filename": filename,
                "original_emotion": img.get("original_emotion", np.nan),
                "text_aoi_count": int(img["text_aoi_count"]),
                "text_present": int(img["text_present"]),
            })

    analysis_df = pd.DataFrame(rows)

    analysis_df = analysis_df.merge(
        responses,
        on=["participant", "filename"],
        how="left"
    )

    analysis_df = analysis_df.dropna(subset=["reported_emotion", "text_present"]).copy()

    if analysis_df.empty:
        raise RuntimeError("No usable rows for text-presence Chi-square analysis.")

    contingency = pd.crosstab(
        analysis_df["reported_emotion"],
        analysis_df["text_present"]
    )

    contingency = contingency.rename(columns={
        0: "Text absent",
        1: "Text present",
    })

    if contingency.shape[0] < 2 or contingency.shape[1] < 2:
        raise RuntimeError(
            "Contingency table does not have at least 2 rows and 2 columns. "
            "This can happen if text is present or absent for all images."
        )

    chi2, p, dof, expected = chi2_contingency(contingency, correction=False)
    v = cramers_v_corrected(contingency.to_numpy())

    expected_arr = np.asarray(expected)

    result = pd.DataFrame([{
        "association": "reported_emotion_x_text_presence",
        "chi_square": chi2,
        "degrees_of_freedom": dof,
        "p_value": p,
        "cramers_v_bias_corrected": v,
        "n": int(contingency.to_numpy().sum()),
        "min_expected_count": float(expected_arr.min()),
        "cells_expected_lt_5": int((expected_arr < 5).sum()),
        "significant_0_05": bool(p < ALPHA),
        "text_present_images": int(images["text_present"].sum()),
        "text_absent_images": int((images["text_present"] == 0).sum()),
    }])

    result.to_csv(RESULT_CSV, index=False, encoding="utf-8-sig")
    contingency.to_csv(CONTINGENCY_CSV, encoding="utf-8-sig")

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Chi-square Text", index=False)
        contingency.to_excel(writer, sheet_name="Contingency")
        analysis_df.to_excel(writer, sheet_name="Analysis Rows", index=False)
        images.to_excel(writer, sheet_name="Image Text Presence", index=False)

    style_excel(XLSX_FILE)

    print("\nDone.")
    print(f"Result:      {RESULT_CSV}")
    print(f"Contingency: {CONTINGENCY_CSV}")
    print(f"Workbook:    {XLSX_FILE}")
    print("\nResult summary:")
    print(result.to_string(index=False))
    print("\nContingency table:")
    print(contingency.to_string())


if __name__ == "__main__":
    main()
