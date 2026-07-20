from pathlib import Path
import random
import warnings

import numpy as np
import pandas as pd
from PIL import Image

import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
from torchvision import models, transforms

from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    accuracy_score,
    precision_recall_fscore_support,
    confusion_matrix,
    classification_report,
)

from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PATHS
# ==========================================================

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

IMAGE_ROOT = BASE_DIR / "imagenes" / "top_imagenes"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"

OUTPUT_DIR = BASE_DIR / "efficientnet_b0_cv"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FOLD_METRICS_CSV = OUTPUT_DIR / "efficientnet_fold_metrics.csv"
PREDICTIONS_CSV = OUTPUT_DIR / "efficientnet_predictions.csv"
CONFUSION_CSV = OUTPUT_DIR / "efficientnet_confusion_matrix.csv"
REPORT_CSV = OUTPUT_DIR / "efficientnet_classification_report.csv"
XLSX_FILE = OUTPUT_DIR / "efficientnet_b0_cv_results.xlsx"


# ==========================================================
# CONFIG
# ==========================================================

SEED = 42
N_SPLITS = 5
EPOCHS = 25
BATCH_SIZE = 8
LEARNING_RATE = 1e-4
WEIGHT_DECAY = 1e-4
PATIENCE = 6

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

# With only 60 images, freeze the backbone by default to reduce overfitting.
FREEZE_BACKBONE = True

# ImageNet normalization.
IMG_SIZE = 224


# ==========================================================
# HELPERS
# ==========================================================

def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def normalize_emotion(value):
    if pd.isna(value):
        return np.nan
    s = str(value).strip()
    sl = s.lower()
    mapping = {
        "positive": "Positive", "positiva": "Positive", "positivo": "Positive",
        "neutral": "Neutral", "neutra": "Neutral", "neutro": "Neutral",
        "negative": "Negative", "negativa": "Negative", "negativo": "Negative",
    }
    return mapping.get(sl, s)


def find_image_files(image_root):
    lookup = {}
    for path in image_root.rglob("*"):
        if path.suffix.lower() in [".jpg", ".jpeg", ".png"]:
            lookup[path.name] = path
    return lookup


def load_dataset_table():
    if not IMAGE_AOI_FILE.exists():
        raise FileNotFoundError(f"Image AOI/order file not found: {IMAGE_AOI_FILE}")

    df = pd.read_csv(IMAGE_AOI_FILE)
    if "emotion" in df.columns and "original_emotion" not in df.columns:
        df = df.rename(columns={"emotion": "original_emotion"})

    required = {"trial_index", "filename", "original_emotion"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{IMAGE_AOI_FILE} missing columns: {sorted(missing)}")

    image_lookup = find_image_files(IMAGE_ROOT)
    df["image_path"] = df["filename"].map(lambda f: str(image_lookup.get(f, "")))
    df["label"] = df["original_emotion"].apply(normalize_emotion)

    df = df.dropna(subset=["label"]).copy()
    df = df[df["image_path"].astype(str) != ""].copy()

    if df.empty:
        raise RuntimeError("No labeled images found.")

    return df.sort_values("trial_index").reset_index(drop=True)


class ImageEmotionDataset(Dataset):
    def __init__(self, df, label_to_idx, transform):
        self.df = df.reset_index(drop=True)
        self.label_to_idx = label_to_idx
        self.transform = transform

    def __len__(self):
        return len(self.df)

    def __getitem__(self, idx):
        row = self.df.iloc[idx]
        image = Image.open(row["image_path"]).convert("RGB")
        image = self.transform(image)
        label = self.label_to_idx[row["label"]]
        return image, label, int(row["trial_index"]), row["filename"]


def get_transforms(train=True):
    if train:
        return transforms.Compose([
            transforms.Resize((IMG_SIZE, IMG_SIZE)),
            transforms.RandomHorizontalFlip(p=0.5),
            transforms.RandomRotation(degrees=8),
            transforms.ColorJitter(brightness=0.15, contrast=0.15, saturation=0.10),
            transforms.ToTensor(),
            transforms.Normalize(
                mean=[0.485, 0.456, 0.406],
                std=[0.229, 0.224, 0.225],
            ),
        ])

    return transforms.Compose([
        transforms.Resize((IMG_SIZE, IMG_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize(
            mean=[0.485, 0.456, 0.406],
            std=[0.229, 0.224, 0.225],
        ),
    ])


def build_model(num_classes):
    weights = models.EfficientNet_B0_Weights.DEFAULT
    model = models.efficientnet_b0(weights=weights)

    if FREEZE_BACKBONE:
        for p in model.features.parameters():
            p.requires_grad = False

    in_features = model.classifier[1].in_features
    model.classifier[1] = nn.Linear(in_features, num_classes)

    return model.to(DEVICE)


def train_one_fold(fold, train_df, val_df, label_to_idx, class_weights):
    train_ds = ImageEmotionDataset(train_df, label_to_idx, get_transforms(train=True))
    val_ds = ImageEmotionDataset(val_df, label_to_idx, get_transforms(train=False))

    train_loader = DataLoader(train_ds, batch_size=BATCH_SIZE, shuffle=True, num_workers=0)
    val_loader = DataLoader(val_ds, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)

    model = build_model(num_classes=len(label_to_idx))

    criterion = nn.CrossEntropyLoss(weight=class_weights.to(DEVICE))
    optimizer = torch.optim.AdamW(
        filter(lambda p: p.requires_grad, model.parameters()),
        lr=LEARNING_RATE,
        weight_decay=WEIGHT_DECAY,
    )

    best_state = None
    best_val_loss = float("inf")
    best_epoch = 0
    epochs_without_improvement = 0

    for epoch in range(1, EPOCHS + 1):
        model.train()
        train_losses = []

        for images, labels, _, _ in train_loader:
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)

            optimizer.zero_grad()
            logits = model(images)
            loss = criterion(logits, labels)
            loss.backward()
            optimizer.step()

            train_losses.append(loss.item())

        model.eval()
        val_losses = []
        with torch.no_grad():
            for images, labels, _, _ in val_loader:
                images = images.to(DEVICE)
                labels = labels.to(DEVICE)
                logits = model(images)
                loss = criterion(logits, labels)
                val_losses.append(loss.item())

        mean_val_loss = float(np.mean(val_losses)) if val_losses else np.nan

        if mean_val_loss < best_val_loss:
            best_val_loss = mean_val_loss
            best_epoch = epoch
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}
            epochs_without_improvement = 0
        else:
            epochs_without_improvement += 1

        if epochs_without_improvement >= PATIENCE:
            break

    if best_state is not None:
        model.load_state_dict(best_state)

    # Evaluate best model.
    model.eval()
    pred_rows = []
    with torch.no_grad():
        for images, labels, trial_indices, filenames in val_loader:
            images = images.to(DEVICE)
            logits = model(images)
            probs = torch.softmax(logits, dim=1).detach().cpu().numpy()
            pred_idx = probs.argmax(axis=1)

            for i in range(len(labels)):
                pred_rows.append({
                    "fold": fold,
                    "trial_index": int(trial_indices[i]),
                    "filename": filenames[i],
                    "true_label": val_df[val_df["filename"] == filenames[i]].iloc[0]["label"],
                    "predicted_label": idx_to_label[int(pred_idx[i])],
                    "predicted_probability": float(probs[i, pred_idx[i]]),
                    **{
                        f"prob_{idx_to_label[j]}": float(probs[i, j])
                        for j in range(len(idx_to_label))
                    },
                })

    preds = pd.DataFrame(pred_rows)
    y_true = preds["true_label"]
    y_pred = preds["predicted_label"]

    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true,
        y_pred,
        average="macro",
        zero_division=0,
    )

    fold_metrics = {
        "fold": fold,
        "n_train": len(train_df),
        "n_val": len(val_df),
        "best_epoch": best_epoch,
        "best_val_loss": best_val_loss,
        "accuracy": accuracy_score(y_true, y_pred),
        "macro_precision": precision,
        "macro_recall": recall,
        "macro_f1": f1,
    }

    return fold_metrics, preds


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)
    wb.save(path)


def main():
    global idx_to_label

    set_seed(SEED)

    print(f"Device: {DEVICE}")
    data = load_dataset_table()

    labels = sorted(data["label"].unique())
    label_to_idx = {label: i for i, label in enumerate(labels)}
    idx_to_label = {i: label for label, i in label_to_idx.items()}

    print("Class counts:")
    print(data["label"].value_counts().to_string())

    y = data["label"].map(label_to_idx).to_numpy()

    min_class_count = data["label"].value_counts().min()
    n_splits = min(N_SPLITS, int(min_class_count))
    if n_splits < 2:
        raise RuntimeError("Not enough samples per class for stratified cross-validation.")

    if n_splits != N_SPLITS:
        warnings.warn(f"Using {n_splits} folds because the smallest class has {min_class_count} samples.")

    # Class weights from full dataset, acceptable for loss weighting but not data leakage in validation labels.
    counts = data["label"].value_counts().reindex(labels).to_numpy()
    weights = counts.sum() / (len(labels) * counts)
    class_weights = torch.tensor(weights, dtype=torch.float32)

    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=SEED)

    all_fold_metrics = []
    all_predictions = []

    for fold, (train_idx, val_idx) in enumerate(skf.split(data, y), start=1):
        print(f"\nFold {fold}/{n_splits}")
        train_df = data.iloc[train_idx].copy()
        val_df = data.iloc[val_idx].copy()

        fold_metrics, preds = train_one_fold(
            fold=fold,
            train_df=train_df,
            val_df=val_df,
            label_to_idx=label_to_idx,
            class_weights=class_weights,
        )

        print(
            f"accuracy={fold_metrics['accuracy']:.3f}, "
            f"macro_f1={fold_metrics['macro_f1']:.3f}, "
            f"best_epoch={fold_metrics['best_epoch']}"
        )

        all_fold_metrics.append(fold_metrics)
        all_predictions.append(preds)

    fold_metrics_df = pd.DataFrame(all_fold_metrics)
    predictions_df = pd.concat(all_predictions, ignore_index=True)

    cm = confusion_matrix(
        predictions_df["true_label"],
        predictions_df["predicted_label"],
        labels=labels,
    )
    cm_df = pd.DataFrame(cm, index=[f"true_{x}" for x in labels], columns=[f"pred_{x}" for x in labels])

    report = classification_report(
        predictions_df["true_label"],
        predictions_df["predicted_label"],
        labels=labels,
        output_dict=True,
        zero_division=0,
    )
    report_df = pd.DataFrame(report).T.reset_index().rename(columns={"index": "class_or_average"})

    overall = pd.DataFrame([{
        "mean_accuracy": fold_metrics_df["accuracy"].mean(),
        "std_accuracy": fold_metrics_df["accuracy"].std(),
        "mean_macro_precision": fold_metrics_df["macro_precision"].mean(),
        "std_macro_precision": fold_metrics_df["macro_precision"].std(),
        "mean_macro_recall": fold_metrics_df["macro_recall"].mean(),
        "std_macro_recall": fold_metrics_df["macro_recall"].std(),
        "mean_macro_f1": fold_metrics_df["macro_f1"].mean(),
        "std_macro_f1": fold_metrics_df["macro_f1"].std(),
        "n_images": len(data),
        "n_folds": n_splits,
        "model": "EfficientNet-B0",
        "pretrained": True,
        "freeze_backbone": FREEZE_BACKBONE,
    }])

    fold_metrics_df.to_csv(FOLD_METRICS_CSV, index=False, encoding="utf-8-sig")
    predictions_df.to_csv(PREDICTIONS_CSV, index=False, encoding="utf-8-sig")
    cm_df.to_csv(CONFUSION_CSV, encoding="utf-8-sig")
    report_df.to_csv(REPORT_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        overall.to_excel(writer, sheet_name="Overall", index=False)
        fold_metrics_df.to_excel(writer, sheet_name="Fold Metrics", index=False)
        predictions_df.to_excel(writer, sheet_name="Predictions", index=False)
        cm_df.to_excel(writer, sheet_name="Confusion Matrix")
        report_df.to_excel(writer, sheet_name="Classification Report", index=False)
        data.to_excel(writer, sheet_name="Dataset", index=False)

    style_excel(XLSX_FILE)

    print("\nDone.")
    print(f"Fold metrics: {FOLD_METRICS_CSV}")
    print(f"Predictions:  {PREDICTIONS_CSV}")
    print(f"Workbook:     {XLSX_FILE}")


if __name__ == "__main__":
    main()
