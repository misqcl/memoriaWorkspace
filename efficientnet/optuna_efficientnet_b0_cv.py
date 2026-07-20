from pathlib import Path
import random
import json
import numpy as np
import pandas as pd
from PIL import Image

import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
from torchvision import models, transforms

from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import accuracy_score, precision_recall_fscore_support, confusion_matrix, classification_report

import optuna
from optuna.samplers import TPESampler

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

IMAGE_ROOT = BASE_DIR / "imagenes" / "top_imagenes"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"

OUTPUT_DIR = BASE_DIR / "efficientnet_b0_optuna_accuracy"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

STUDY_DB = OUTPUT_DIR / "optuna_study.db"
TRIALS_CSV = OUTPUT_DIR / "optuna_trials.csv"
BEST_PREDICTIONS_CSV = OUTPUT_DIR / "best_trial_predictions.csv"
BEST_FOLD_METRICS_CSV = OUTPUT_DIR / "best_trial_fold_metrics.csv"
CONFUSION_CSV = OUTPUT_DIR / "best_trial_confusion_matrix.csv"
REPORT_CSV = OUTPUT_DIR / "best_trial_classification_report.csv"
XLSX_FILE = OUTPUT_DIR / "optuna_efficientnet_b0_results.xlsx"

SEED = 42
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
IMG_SIZE = 224
NUM_WORKERS = 0

# Runtime controls. Increase N_TRIALS if you have time/GPU.
N_TRIALS = 20
MAX_EPOCHS = 18
PATIENCE = 4
MIN_FOLDS = 3
MAX_FOLDS = 8
OBJECTIVE_METRIC = "accuracy"  # or "accuracy"


def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def normalize_emotion(value):
    if pd.isna(value):
        return np.nan
    s = str(value).strip().lower()
    mapping = {
        "positive": "Positive", "positiva": "Positive", "positivo": "Positive",
        "neutral": "Neutral", "neutra": "Neutral", "neutro": "Neutral",
        "negative": "Negative", "negativa": "Negative", "negativo": "Negative",
    }
    return mapping.get(s, str(value).strip())


def find_image_files(image_root):
    lookup = {}
    for path in image_root.rglob("*"):
        if path.suffix.lower() in [".jpg", ".jpeg", ".png"]:
            lookup[path.name] = path
    return lookup


def load_dataset_table():
    df = pd.read_csv(IMAGE_AOI_FILE)
    if "emotion" in df.columns and "original_emotion" not in df.columns:
        df = df.rename(columns={"emotion": "original_emotion"})

    required = {"trial_index", "filename", "original_emotion"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{IMAGE_AOI_FILE} missing columns: {sorted(missing)}")

    image_lookup = find_image_files(IMAGE_ROOT)
    df["image_path"] = df["filename"].map(lambda f: str(image_lookup.get(f, "")))
    df["label"] = df["original_emotion"].apply(normalize_emotion)
    df = df.dropna(subset=["label"]).copy()
    df = df[df["image_path"].astype(str) != ""].copy()
    df = df.sort_values("trial_index").reset_index(drop=True)

    if df.empty:
        raise RuntimeError("No labeled images found.")
    return df


class ImageEmotionDataset(Dataset):
    def __init__(self, df, label_to_idx, transform):
        self.df = df.reset_index(drop=True)
        self.label_to_idx = label_to_idx
        self.transform = transform

    def __len__(self):
        return len(self.df)

    def __getitem__(self, idx):
        row = self.df.iloc[idx]
        image = Image.open(row["image_path"]).convert("RGB")
        image = self.transform(image)
        label = self.label_to_idx[row["label"]]
        return image, label, int(row["trial_index"]), row["filename"]


def get_transforms(params, train=True):
    if train:
        ops = [
            transforms.Resize((IMG_SIZE, IMG_SIZE)),
            transforms.RandomHorizontalFlip(p=params["horizontal_flip_p"]),
        ]
        if params["rotation_degrees"] > 0:
            ops.append(transforms.RandomRotation(degrees=params["rotation_degrees"]))
        if params["color_jitter"]:
            ops.append(transforms.ColorJitter(
                brightness=params["brightness"],
                contrast=params["contrast"],
                saturation=params["saturation"],
            ))
        ops += [
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
        ]
        return transforms.Compose(ops)

    return transforms.Compose([
        transforms.Resize((IMG_SIZE, IMG_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])


def build_model(num_classes, freeze_backbone, dropout_p):
    weights = models.EfficientNet_B0_Weights.DEFAULT
    model = models.efficientnet_b0(weights=weights)

    if freeze_backbone:
        for p in model.features.parameters():
            p.requires_grad = False

    in_features = model.classifier[1].in_features
    model.classifier = nn.Sequential(nn.Dropout(p=dropout_p), nn.Linear(in_features, num_classes))
    return model.to(DEVICE)


def make_class_weights(train_df, labels):
    counts = train_df["label"].value_counts().reindex(labels).fillna(0).to_numpy(dtype=float)
    counts = np.maximum(counts, 1)
    weights = counts.sum() / (len(labels) * counts)
    return torch.tensor(weights, dtype=torch.float32)


def train_one_fold(fold, train_df, val_df, label_to_idx, idx_to_label, labels, params, return_predictions=False):
    train_ds = ImageEmotionDataset(train_df, label_to_idx, get_transforms(params, train=True))
    val_ds = ImageEmotionDataset(val_df, label_to_idx, get_transforms(params, train=False))

    train_loader = DataLoader(train_ds, batch_size=params["batch_size"], shuffle=True, num_workers=NUM_WORKERS)
    val_loader = DataLoader(val_ds, batch_size=params["batch_size"], shuffle=False, num_workers=NUM_WORKERS)

    model = build_model(len(label_to_idx), params["freeze_backbone"], params["dropout_p"])
    criterion = nn.CrossEntropyLoss(weight=make_class_weights(train_df, labels).to(DEVICE))

    optim_cls = torch.optim.AdamW if params["optimizer"] == "AdamW" else torch.optim.Adam
    optimizer = optim_cls(
        filter(lambda p: p.requires_grad, model.parameters()),
        lr=params["lr"],
        weight_decay=params["weight_decay"],
    )

    best_state = None
    best_val_loss = float("inf")
    best_epoch = 0
    no_improve = 0

    for epoch in range(1, MAX_EPOCHS + 1):
        model.train()
        for images, y, _, _ in train_loader:
            images = images.to(DEVICE)
            y = y.to(DEVICE)
            optimizer.zero_grad()
            loss = criterion(model(images), y)
            loss.backward()
            optimizer.step()

        model.eval()
        val_losses = []
        with torch.no_grad():
            for images, y, _, _ in val_loader:
                images = images.to(DEVICE)
                y = y.to(DEVICE)
                val_losses.append(float(criterion(model(images), y).item()))

        val_loss = float(np.mean(val_losses)) if val_losses else np.nan
        if val_loss < best_val_loss:
            best_val_loss = val_loss
            best_epoch = epoch
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}
            no_improve = 0
        else:
            no_improve += 1
        if no_improve >= PATIENCE:
            break

    if best_state is not None:
        model.load_state_dict(best_state)

    rows = []
    model.eval()
    with torch.no_grad():
        for images, y, trial_indices, filenames in val_loader:
            probs = torch.softmax(model(images.to(DEVICE)), dim=1).detach().cpu().numpy()
            pred_idx = probs.argmax(axis=1)
            for i, filename in enumerate(filenames):
                true_label = val_df[val_df["filename"] == filename].iloc[0]["label"]
                row = {
                    "fold": fold,
                    "trial_index": int(trial_indices[i]),
                    "filename": filename,
                    "true_label": true_label,
                    "predicted_label": idx_to_label[int(pred_idx[i])],
                    "predicted_probability": float(probs[i, pred_idx[i]]),
                }
                for j in range(len(idx_to_label)):
                    row[f"prob_{idx_to_label[j]}"] = float(probs[i, j])
                rows.append(row)

    preds = pd.DataFrame(rows)
    precision, recall, f1, _ = precision_recall_fscore_support(
        preds["true_label"], preds["predicted_label"], average="macro", zero_division=0
    )
    metrics = {
        "fold": fold,
        "n_train": len(train_df),
        "n_val": len(val_df),
        "best_epoch": best_epoch,
        "best_val_loss": best_val_loss,
        "accuracy": accuracy_score(preds["true_label"], preds["predicted_label"]),
        "macro_precision": precision,
        "macro_recall": recall,
        "macro_f1": f1,
    }
    return (metrics, preds) if return_predictions else (metrics, None)


def run_cv(data, label_to_idx, idx_to_label, params, return_predictions=False):
    labels = sorted(label_to_idx.keys())
    y = data["label"].map(label_to_idx).to_numpy()
    skf = StratifiedKFold(n_splits=params["n_splits"], shuffle=True, random_state=SEED)
    metric_rows, pred_rows = [], []

    for fold, (train_idx, val_idx) in enumerate(skf.split(data, y), start=1):
        metrics, preds = train_one_fold(
            fold, data.iloc[train_idx].copy(), data.iloc[val_idx].copy(),
            label_to_idx, idx_to_label, labels, params, return_predictions=return_predictions
        )
        metric_rows.append(metrics)
        if return_predictions:
            pred_rows.append(preds)

    metrics_df = pd.DataFrame(metric_rows)
    preds_df = pd.concat(pred_rows, ignore_index=True) if pred_rows else pd.DataFrame()
    overall = {
        "mean_accuracy": float(metrics_df["accuracy"].mean()),
        "std_accuracy": float(metrics_df["accuracy"].std()),
        "mean_macro_precision": float(metrics_df["macro_precision"].mean()),
        "std_macro_precision": float(metrics_df["macro_precision"].std()),
        "mean_macro_recall": float(metrics_df["macro_recall"].mean()),
        "std_macro_recall": float(metrics_df["macro_recall"].std()),
        "mean_macro_f1": float(metrics_df["macro_f1"].mean()),
        "std_macro_f1": float(metrics_df["macro_f1"].std()),
        "mean_best_epoch": float(metrics_df["best_epoch"].mean()),
        "mean_best_val_loss": float(metrics_df["best_val_loss"].mean()),
    }
    return overall, metrics_df, preds_df


def suggest_params(trial, max_allowed_folds):
    params = {
        "n_splits": trial.suggest_int("n_splits", MIN_FOLDS, min(MAX_FOLDS, max_allowed_folds)),
        "batch_size": trial.suggest_categorical("batch_size", [4, 8, 12]),
        "lr": trial.suggest_float("lr", 1e-5, 8e-4, log=True),
        "weight_decay": trial.suggest_float("weight_decay", 1e-6, 5e-3, log=True),
        "dropout_p": trial.suggest_float("dropout_p", 0.1, 0.6),
        "freeze_backbone": trial.suggest_categorical("freeze_backbone", [True, False]),
        "optimizer": trial.suggest_categorical("optimizer", ["AdamW", "Adam"]),
        "horizontal_flip_p": trial.suggest_float("horizontal_flip_p", 0.0, 0.5),
        "rotation_degrees": trial.suggest_int("rotation_degrees", 0, 15),
        "color_jitter": trial.suggest_categorical("color_jitter", [True, False]),
        "brightness": trial.suggest_float("brightness", 0.0, 0.25),
        "contrast": trial.suggest_float("contrast", 0.0, 0.25),
        "saturation": trial.suggest_float("saturation", 0.0, 0.20),
    }
    if not params["color_jitter"]:
        params["brightness"] = params["contrast"] = params["saturation"] = 0.0
    return params


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)
    wb.save(path)


def main():
    set_seed(SEED)
    print(f"Device: {DEVICE}")
    print(f"N_TRIALS={N_TRIALS}, MAX_EPOCHS={MAX_EPOCHS}, PATIENCE={PATIENCE}")

    data = load_dataset_table()
    labels = sorted(data["label"].unique())
    label_to_idx = {label: i for i, label in enumerate(labels)}
    idx_to_label = {i: label for label, i in label_to_idx.items()}

    print("Class counts:")
    print(data["label"].value_counts().to_string())

    min_class_count = int(data["label"].value_counts().min())
    max_allowed_folds = min(MAX_FOLDS, min_class_count)
    if max_allowed_folds < MIN_FOLDS:
        raise RuntimeError(f"Cannot run at least {MIN_FOLDS} folds; smallest class has {min_class_count} images.")

    def objective(trial):
        set_seed(SEED + trial.number)
        params = suggest_params(trial, max_allowed_folds)
        overall, _, _ = run_cv(data, label_to_idx, idx_to_label, params, return_predictions=False)
        for k, v in overall.items():
            trial.set_user_attr(k, v)
        return overall[f"mean_{OBJECTIVE_METRIC}"]

    study = optuna.create_study(
        direction="maximize",
        sampler=TPESampler(seed=SEED),
        study_name="efficientnet_b0_cv_optuna",
        storage=f"sqlite:///{STUDY_DB}",
        load_if_exists=True,
    )
    study.optimize(objective, n_trials=N_TRIALS)

    trials_df = study.trials_dataframe(attrs=("number", "value", "params", "user_attrs", "state"))
    trials_df.to_csv(TRIALS_CSV, index=False, encoding="utf-8-sig")

    best = dict(study.best_trial.params)
    best_params = {
        "n_splits": int(best["n_splits"]),
        "batch_size": int(best["batch_size"]),
        "lr": float(best["lr"]),
        "weight_decay": float(best["weight_decay"]),
        "dropout_p": float(best["dropout_p"]),
        "freeze_backbone": bool(best["freeze_backbone"]),
        "optimizer": best["optimizer"],
        "horizontal_flip_p": float(best["horizontal_flip_p"]),
        "rotation_degrees": int(best["rotation_degrees"]),
        "color_jitter": bool(best["color_jitter"]),
        "brightness": float(best.get("brightness", 0.0)),
        "contrast": float(best.get("contrast", 0.0)),
        "saturation": float(best.get("saturation", 0.0)),
    }
    if not best_params["color_jitter"]:
        best_params["brightness"] = best_params["contrast"] = best_params["saturation"] = 0.0

    print("\nBest trial:")
    print(f"Value ({OBJECTIVE_METRIC}): {study.best_value:.4f}")
    print(json.dumps(best_params, indent=2))

    print("\nRerunning best configuration to save predictions...")
    set_seed(SEED)
    overall, fold_metrics, predictions = run_cv(data, label_to_idx, idx_to_label, best_params, return_predictions=True)

    fold_metrics.to_csv(BEST_FOLD_METRICS_CSV, index=False, encoding="utf-8-sig")
    predictions.to_csv(BEST_PREDICTIONS_CSV, index=False, encoding="utf-8-sig")

    cm = confusion_matrix(predictions["true_label"], predictions["predicted_label"], labels=labels)
    cm_df = pd.DataFrame(cm, index=[f"true_{x}" for x in labels], columns=[f"pred_{x}" for x in labels])
    cm_df.to_csv(CONFUSION_CSV, encoding="utf-8-sig")

    report_df = pd.DataFrame(classification_report(
        predictions["true_label"], predictions["predicted_label"], labels=labels, output_dict=True, zero_division=0
    )).T.reset_index().rename(columns={"index": "class_or_average"})
    report_df.to_csv(REPORT_CSV, index=False, encoding="utf-8-sig")

    best_overall = pd.DataFrame([{**overall, **{f"best_{k}": v for k, v in best_params.items()},
                                  "objective_metric": OBJECTIVE_METRIC,
                                  "optuna_best_value": study.best_value,
                                  "n_trials_total_in_db": len(study.trials),
                                  "device": DEVICE,
                                  "max_epochs": MAX_EPOCHS,
                                  "patience": PATIENCE}])
    params_df = pd.DataFrame([best_params])

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        best_overall.to_excel(writer, sheet_name="Best Overall", index=False)
        params_df.to_excel(writer, sheet_name="Best Params", index=False)
        fold_metrics.to_excel(writer, sheet_name="Best Fold Metrics", index=False)
        predictions.to_excel(writer, sheet_name="Best Predictions", index=False)
        cm_df.to_excel(writer, sheet_name="Best Confusion Matrix")
        report_df.to_excel(writer, sheet_name="Best Class Report", index=False)
        trials_df.to_excel(writer, sheet_name="Optuna Trials", index=False)
        data.to_excel(writer, sheet_name="Dataset", index=False)
    style_excel(XLSX_FILE)

    print("\nDone.")
    print(f"Trials:            {TRIALS_CSV}")
    print(f"Best predictions:  {BEST_PREDICTIONS_CSV}")
    print(f"Best fold metrics: {BEST_FOLD_METRICS_CSV}")
    print(f"Workbook:          {XLSX_FILE}")
    print(f"Study DB:          {STUDY_DB}")


if __name__ == "__main__":
    main()
