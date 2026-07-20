from pathlib import Path
import warnings

import numpy as np
import pandas as pd
from PIL import Image
from scipy.special import logsumexp
from scipy.ndimage import gaussian_filter
import imageio.v2 as imageio

import torch
import matplotlib.pyplot as plt
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PATHS
# ==========================================================

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

IMAGE_ROOT = BASE_DIR / "imagenes" / "top_imagenes"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"

OUTPUT_DIR = BASE_DIR / "deepgaze_predictions"
RAW_DIR = OUTPUT_DIR / "raw"
LOG_DIR = OUTPUT_DIR / "log_density"
PREVIEW_DIR = OUTPUT_DIR / "previews"
OVERLAY_DIR = OUTPUT_DIR / "overlay_previews"

SUMMARY_CSV = OUTPUT_DIR / "deepgaze_prediction_summary.csv"
SUMMARY_XLSX = OUTPUT_DIR / "deepgaze_prediction_summary.xlsx"


# ==========================================================
# CONFIG
# ==========================================================

MODEL_NAME = "DeepGazeIIE"  # Best default for this thesis step.
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

CENTERBIAS_SIGMA_FRACTION = 0.25

SAVE_PREVIEWS = True


# ==========================================================
# HELPERS
# ==========================================================

def ensure_dirs():
    for d in [OUTPUT_DIR, RAW_DIR, LOG_DIR, PREVIEW_DIR, OVERLAY_DIR]:
        d.mkdir(parents=True, exist_ok=True)


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)
    wb.save(path)


def find_image_files(image_root):
    lookup = {}
    for path in image_root.rglob("*"):
        if path.suffix.lower() in [".jpg", ".jpeg", ".png"]:
            lookup[path.name] = path
    return lookup


def load_image_order():
    if not IMAGE_AOI_FILE.exists():
        raise FileNotFoundError(f"Image AOI/order file not found: {IMAGE_AOI_FILE}")

    df = pd.read_csv(IMAGE_AOI_FILE)
    if "emotion" in df.columns and "original_emotion" not in df.columns:
        df = df.rename(columns={"emotion": "original_emotion"})

    required = {"trial_index", "filename"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{IMAGE_AOI_FILE} missing columns: {sorted(missing)}")

    return df.sort_values("trial_index").drop_duplicates("filename")


def make_centerbias_log_density(height, width):
    yy, xx = np.mgrid[0:height, 0:width]
    cy = (height - 1) / 2.0
    cx = (width - 1) / 2.0
    sigma = CENTERBIAS_SIGMA_FRACTION * min(height, width)

    if sigma <= 0:
        centerbias = np.zeros((height, width), dtype=np.float32)
    else:
        centerbias = -(((xx - cx) ** 2 + (yy - cy) ** 2) / (2 * sigma ** 2))

    centerbias = centerbias.astype(np.float64)
    centerbias -= logsumexp(centerbias)
    return centerbias.astype(np.float32)


def normalize_density_from_log(log_density):
    log_density = np.asarray(log_density, dtype=np.float64)
    log_density = log_density - logsumexp(log_density)
    density = np.exp(log_density)
    density = density / density.sum()
    return density.astype(np.float32), log_density.astype(np.float32)


def load_deepgaze_model():
    try:
        import deepgaze_pytorch
    except Exception as exc:
        raise ImportError(
            "Could not import deepgaze_pytorch. Install it inside the same Python "
            "environment used to run this script. Try:\n"
            "  python -m pip install deepgaze-pytorch\n"
            f"Original import error: {exc}"
        )

    if not hasattr(deepgaze_pytorch, MODEL_NAME):
        available = [x for x in dir(deepgaze_pytorch) if "DeepGaze" in x]
        raise AttributeError(
            f"deepgaze_pytorch does not expose {MODEL_NAME}. "
            f"Available DeepGaze-like names: {available}"
        )

    model_cls = getattr(deepgaze_pytorch, MODEL_NAME)

    try:
        model = model_cls(pretrained=True)
    except TypeError:
        model = model_cls()

    model = model.to(DEVICE)
    model.eval()
    return model


def predict_log_density(model, image_rgb):
    h, w = image_rgb.shape[:2]

    image_tensor = torch.tensor(
        image_rgb.transpose(2, 0, 1)[None, ...],
        dtype=torch.float32,
        device=DEVICE,
    )

    centerbias = make_centerbias_log_density(h, w)
    centerbias_tensor = torch.tensor(
        centerbias[None, ...],
        dtype=torch.float32,
        device=DEVICE,
    )

    with torch.no_grad():
        output = model(image_tensor, centerbias_tensor)

    out = output.detach().cpu().numpy()

    out = np.squeeze(out)

    if out.ndim != 2:
        raise ValueError(f"Unexpected DeepGaze output shape after squeeze: {out.shape}")

    return out


def save_preview_images(image_rgb, density, preview_path, overlay_path):
    # Saliency heatmap only.
    plt.figure(figsize=(6, 6))
    plt.imshow(density, cmap="inferno")
    plt.axis("off")
    plt.tight_layout(pad=0)
    plt.savefig(preview_path, dpi=150, bbox_inches="tight", pad_inches=0)
    plt.close()

    # Overlay.
    plt.figure(figsize=(6, 6))
    plt.imshow(image_rgb)
    plt.imshow(density, cmap="inferno", alpha=0.45)
    plt.axis("off")
    plt.tight_layout(pad=0)
    plt.savefig(overlay_path, dpi=150, bbox_inches="tight", pad_inches=0)
    plt.close()


def main():
    ensure_dirs()

    print(f"Device: {DEVICE}")
    print(f"Loading {MODEL_NAME}...")
    model = load_deepgaze_model()

    image_order = load_image_order()
    image_lookup = find_image_files(IMAGE_ROOT)

    summary_rows = []

    for _, row in tqdm(image_order.iterrows(), total=len(image_order), desc="DeepGaze"):
        trial_index = int(row["trial_index"])
        filename = row["filename"]

        if filename not in image_lookup:
            warnings.warn(f"Missing image: {filename}")
            summary_rows.append({
                "trial_index": trial_index,
                "filename": filename,
                "status": "missing_image",
            })
            continue

        image_path = image_lookup[filename]

        try:
            image = Image.open(image_path).convert("RGB")
            image_rgb = np.asarray(image)
        except Exception as exc:
            warnings.warn(f"Could not read {filename}: {exc}")
            summary_rows.append({
                "trial_index": trial_index,
                "filename": filename,
                "status": "read_error",
            })
            continue

        h, w = image_rgb.shape[:2]

        try:
            log_density = predict_log_density(model, image_rgb)
            density, log_density_norm = normalize_density_from_log(log_density)
        except Exception as exc:
            warnings.warn(f"Prediction failed for {filename}: {exc}")
            summary_rows.append({
                "trial_index": trial_index,
                "filename": filename,
                "status": "prediction_error",
                "error": str(exc),
            })
            continue

        stem = Path(filename).stem
        raw_path = RAW_DIR / f"{trial_index:02d}_{stem}_deepgaze_density.npy"
        log_path = LOG_DIR / f"{trial_index:02d}_{stem}_deepgaze_log_density.npy"
        preview_path = PREVIEW_DIR / f"{trial_index:02d}_{stem}_deepgaze.png"
        overlay_path = OVERLAY_DIR / f"{trial_index:02d}_{stem}_deepgaze_overlay.png"

        np.save(raw_path, density)
        np.save(log_path, log_density_norm)

        if SAVE_PREVIEWS:
            save_preview_images(image_rgb, density, preview_path, overlay_path)

        summary_rows.append({
            "trial_index": trial_index,
            "filename": filename,
            "status": "ok",
            "model": MODEL_NAME,
            "device": DEVICE,
            "img_w": w,
            "img_h": h,
            "density_sum": float(density.sum()),
            "density_min": float(density.min()),
            "density_max": float(density.max()),
            "raw_density_path": str(raw_path),
            "log_density_path": str(log_path),
            "preview_path": str(preview_path),
            "overlay_path": str(overlay_path),
        })

    summary = pd.DataFrame(summary_rows)
    summary.to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(SUMMARY_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="DeepGaze Predictions", index=False)

    style_excel(SUMMARY_XLSX)

    print("\nDone.")
    print(f"Summary CSV:  {SUMMARY_CSV}")
    print(f"Summary XLSX: {SUMMARY_XLSX}")
    print(f"Raw maps:     {RAW_DIR}")
    print(f"Previews:     {PREVIEW_DIR}")
    print(f"Overlays:     {OVERLAY_DIR}")


if __name__ == "__main__":
    main()
