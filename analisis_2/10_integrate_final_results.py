from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from analysis_2_common import normalize_emotion as normalize_emotion_no_nr


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"
ANALYSIS_DIR = BASE_DIR / "analisis_2"
HEATMAP_XLSX = ANALYSIS_DIR / "heatmap_analysis" / "heatmap_analysis_clean_no_nr.xlsx"
AOI_WIDE_CSV = ANALYSIS_DIR / "aoi_analysis_EOCR" / "aoi_metrics_wide.csv"
DEEPGAZE_CSV = ANALYSIS_DIR / "deepgaze_evaluation" / "deepgaze_metrics_by_image.csv"
EFFICIENTNET_PRED_CSV = BASE_DIR / "efficientnet_b0_optuna" / "best_trial_predictions.csv"

OUTPUT_DIR = ANALYSIS_DIR / "final_integrated_interpretation"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_XLSX = OUTPUT_DIR / "final_integrated_interpretation.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "final_integrated_image_level.csv"


def normalize_emotion(value):
    if pd.isna(value):
        return np.nan
    s = str(value).strip()
    sl = s.lower()
    mapping = {
        "positive": "Positive", "positiva": "Positive", "positivo": "Positive",
        "neutral": "Neutral", "neutra": "Neutral", "neutro": "Neutral",
        "negative": "Negative", "negativa": "Negative", "negativo": "Negative",
        "no responde": "Neutral",
        "no response": "Neutral",
    }
    return mapping.get(sl, s)


# Option B override: NR values are excluded rather than recoded as Neutral.
normalize_emotion = normalize_emotion_no_nr

def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)
    wb.save(path)


def load_base_images():
    df = pd.read_csv(IMAGE_AOI_FILE)
    if "emotion" in df.columns and "original_emotion" not in df.columns:
        df = df.rename(columns={"emotion": "original_emotion"})
    df["original_emotion"] = df["original_emotion"].apply(normalize_emotion)
    return df[["trial_index", "filename", "original_emotion"]].drop_duplicates("filename")


def summarize_heatmap_behavior():
    if not HEATMAP_XLSX.exists():
        return pd.DataFrame()

    xls = pd.ExcelFile(HEATMAP_XLSX)
    # Prefer the most granular sheet if present.
    preferred = None
    for s in xls.sheet_names:
        low = s.lower()
        if "image" in low and ("user" in low or "participant" in low):
            preferred = s
            break
    if preferred is None:
        preferred = xls.sheet_names[0]

    df = pd.read_excel(HEATMAP_XLSX, sheet_name=preferred)

    if "filename" not in df.columns:
        return pd.DataFrame()

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    wanted = [
        c for c in numeric_cols
        if any(k in c.lower() for k in ["fixation", "dwell", "duration", "density"])
    ]

    if not wanted:
        return pd.DataFrame()

    agg = df.groupby("filename", as_index=False)[wanted].mean()
    agg = agg.rename(columns={c: f"viewing_mean_{c}" for c in wanted if c != "filename"})
    return agg


def summarize_aoi_behavior():
    if not AOI_WIDE_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(AOI_WIDE_CSV)
    if "filename" not in df.columns:
        return pd.DataFrame()

    metric_cols = [
        c for c in df.select_dtypes(include=[np.number]).columns
        if any(k in c for k in [
            "face_fixation_proportion", "text_fixation_proportion",
            "face_dwell_proportion", "text_dwell_proportion",
            "face_area_normalized_density", "text_area_normalized_density",
            "face_ttff_ms", "text_ttff_ms"
        ])
    ]

    if not metric_cols:
        metric_cols = [
            c for c in df.select_dtypes(include=[np.number]).columns
            if c not in ["trial_index"]
        ]

    agg = df.groupby("filename", as_index=False)[metric_cols].mean()
    agg = agg.rename(columns={c: f"aoi_mean_{c}" for c in metric_cols})
    return agg


def summarize_perceived_emotion():
    if not AOI_WIDE_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(AOI_WIDE_CSV)
    if not {"filename", "reported_emotion"}.issubset(df.columns):
        return pd.DataFrame()

    df["reported_emotion"] = df["reported_emotion"].apply(normalize_emotion)
    rows = []
    for filename, g in df.dropna(subset=["reported_emotion"]).groupby("filename"):
        counts = g["reported_emotion"].value_counts()
        total = counts.sum()
        majority = counts.index[0] if total else np.nan
        row = {
            "filename": filename,
            "perceived_emotion_majority": majority,
            "perceived_emotion_total_responses": int(total),
        }
        for label in ["Positive", "Neutral", "Negative"]:
            row[f"perceived_prop_{label}"] = counts.get(label, 0) / total if total else np.nan
            row[f"perceived_count_{label}"] = int(counts.get(label, 0))
        rows.append(row)

    return pd.DataFrame(rows)


def summarize_deepgaze():
    if not DEEPGAZE_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(DEEPGAZE_CSV)
    keep = [
        c for c in [
            "filename", "CC", "SIM", "KL_empirical_to_prediction", "AUC_Judd", "NSS",
            "n_fixations_for_auc_nss"
        ]
        if c in df.columns
    ]
    return df[keep].drop_duplicates("filename")


def summarize_efficientnet():
    if not EFFICIENTNET_PRED_CSV.exists():
        return pd.DataFrame()

    df = pd.read_csv(EFFICIENTNET_PRED_CSV)
    keep = [
        c for c in [
            "filename", "true_label", "predicted_label", "predicted_probability"
        ]
        if c in df.columns
    ]
    out = df[keep].drop_duplicates("filename").copy()
    out = out.rename(columns={
        "true_label": "efficientnet_true_label",
        "predicted_label": "efficientnet_predicted_label",
        "predicted_probability": "efficientnet_predicted_probability",
    })
    return out


def add_interpretive_flags(df):
    if {"original_emotion", "perceived_emotion_majority"}.issubset(df.columns):
        df["intended_matches_perceived_majority"] = (
            df["original_emotion"].apply(normalize_emotion) ==
            df["perceived_emotion_majority"].apply(normalize_emotion)
        )

    if {"original_emotion", "efficientnet_predicted_label"}.issubset(df.columns):
        df["intended_matches_efficientnet"] = (
            df["original_emotion"].apply(normalize_emotion) ==
            df["efficientnet_predicted_label"].apply(normalize_emotion)
        )

    # Simple DeepGaze interpretation buckets.
    if "NSS" in df.columns:
        df["deepgaze_nss_agreement_level"] = pd.cut(
            df["NSS"],
            bins=[-np.inf, 0, 1, 2, np.inf],
            labels=["poor_or_inverse", "weak", "moderate", "strong"]
        )

    if "AUC_Judd" in df.columns:
        df["deepgaze_auc_agreement_level"] = pd.cut(
            df["AUC_Judd"],
            bins=[-np.inf, 0.5, 0.6, 0.7, np.inf],
            labels=["chance_or_worse", "weak", "moderate", "strong"]
        )

    # Semantic AOI emphasis: which semantic region drew more proportional attention.
    face_col = next((c for c in df.columns if c.endswith("face_fixation_proportion")), None)
    text_col = next((c for c in df.columns if c.endswith("text_fixation_proportion")), None)
    if face_col and text_col:
        conditions = [
            df[face_col] > df[text_col],
            df[text_col] > df[face_col],
            df[face_col].eq(df[text_col]),
        ]
        choices = ["face_more_than_text", "text_more_than_face", "equal_face_text"]
        df["semantic_attention_emphasis"] = np.select(
            conditions,
            choices,
            default="unknown"
        )

    return df


def main():
    base = load_base_images()

    tables = [
        summarize_perceived_emotion(),
        summarize_heatmap_behavior(),
        summarize_aoi_behavior(),
        summarize_deepgaze(),
        summarize_efficientnet(),
    ]

    integrated = base.copy()
    for t in tables:
        if not t.empty and "filename" in t.columns:
            integrated = integrated.merge(t, on="filename", how="left")

    integrated = add_interpretive_flags(integrated)
    integrated.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    # Summary tables.
    emotion_summary_cols = [
        c for c in [
            "original_emotion", "CC", "SIM", "KL_empirical_to_prediction", "AUC_Judd", "NSS"
        ]
        if c in integrated.columns
    ]
    if len(emotion_summary_cols) > 1:
        by_emotion_deepgaze = (
            integrated.groupby("original_emotion")[emotion_summary_cols[1:]]
            .agg(["count", "mean", "std", "median"])
            .reset_index()
        )
        by_emotion_deepgaze.columns = [
            "_".join([str(x) for x in col if str(x)])
            if isinstance(col, tuple) else str(col)
            for col in by_emotion_deepgaze.columns
        ]
    else:
        by_emotion_deepgaze = pd.DataFrame()

    if {"original_emotion", "perceived_emotion_majority"}.issubset(integrated.columns):
        intended_perceived = pd.crosstab(
            integrated["original_emotion"],
            integrated["perceived_emotion_majority"],
            margins=True
        )
    else:
        intended_perceived = pd.DataFrame()

    if {"original_emotion", "efficientnet_predicted_label"}.issubset(integrated.columns):
        intended_model = pd.crosstab(
            integrated["original_emotion"],
            integrated["efficientnet_predicted_label"],
            margins=True
        )
    else:
        intended_model = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        integrated.to_excel(writer, sheet_name="Integrated Image Level", index=False)
        if not by_emotion_deepgaze.empty:
            by_emotion_deepgaze.to_excel(writer, sheet_name="DeepGaze by Emotion", index=False)
        if not intended_perceived.empty:
            intended_perceived.to_excel(writer, sheet_name="Intended vs Perceived")
        if not intended_model.empty:
            intended_model.to_excel(writer, sheet_name="Intended vs EfficientNet")

    style_excel(OUTPUT_XLSX)

    print("Done.")
    print(f"Integrated CSV:  {OUTPUT_CSV}")
    print(f"Integrated XLSX: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
