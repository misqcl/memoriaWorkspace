# Install: pip install pandas numpy scipy statsmodels openpyxl

from pathlib import Path
from itertools import combinations
import re
import numpy as np
import pandas as pd
from scipy.stats import shapiro, levene, f_oneway, kruskal, rankdata, norm, mannwhitneyu
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from analysis_2_common import is_no_response, normalize_emotion as normalize_emotion_no_nr

SCRIPT_DIR=Path(__file__).resolve().parent
BASE_DIR=SCRIPT_DIR.parent
RESULTS_DIR=BASE_DIR/"results"
AOI_FILE=BASE_DIR/"image_aois"/"image_aoi_seed42.csv"
QA_FILE=BASE_DIR/"fixation_summary_QA.xlsx"
TODOS_DIR=BASE_DIR/"Todos"
SUMMARY_FILE=TODOS_DIR/"Summary.xlsx"
ANALYSIS_DIR=BASE_DIR/"analisis_2"
HEATMAP_DIR=ANALYSIS_DIR/"image_heatmaps"/"raw"
OUTPUT_DIR=ANALYSIS_DIR/"heatmap_analysis"; OUTPUT_DIR.mkdir(parents=True,exist_ok=True)
OUTPUT_FILE=OUTPUT_DIR/"heatmap_analysis_clean_no_nr.xlsx"
QA_SHEET="Usable"; N_IMAGES=60; N_USERS=38; ALPHA=.05

def normalize_emotion(v):
    if pd.isna(v): return np.nan
    s=str(v).strip().upper()
    m={"POSITIVO":"Positivo","POSITIVA":"Positivo","POSITIVE":"Positivo",
       "NEUTRAL":"Neutral","NEGATIVO":"Negativo","NEGATIVA":"Negativo",
       "NEGATIVE":"Negativo","NO RESPONDE":"Neutral","NO RESPONSE":"Neutral",
       "NO_RESPONSE":"Neutral"}
    return m.get(s,str(v).strip())

# NR/No responde are excluded, never recoded as Neutral.
normalize_emotion = normalize_emotion_no_nr

def clean_name(v):
    return "" if pd.isna(v) else re.sub(r"\s+","",str(v).strip())

def safe_shapiro(v):
    v=pd.Series(v).dropna()
    if len(v)<3:return np.nan
    if len(v)>5000:v=v.sample(5000,random_state=42)
    try:return shapiro(v).pvalue
    except:return np.nan

def safe_levene(groups):
    g=[pd.Series(x).dropna() for x in groups if len(pd.Series(x).dropna())>=2]
    if len(g)<2:return np.nan
    try:return levene(*g).pvalue
    except:return np.nan

def eta_squared(groups):
    a=[np.asarray(g,float) for g in groups]
    allv=np.concatenate(a); gm=allv.mean()
    ssb=sum(len(g)*(g.mean()-gm)**2 for g in a)
    sst=((allv-gm)**2).sum()
    return ssb/sst if sst else np.nan

def epsilon_squared(h,n,k):
    return max(0.,(h-k+1)/(n-k)) if n>k else np.nan

def rank_biserial(x,y):
    x=np.asarray(pd.Series(x).dropna(),float); y=np.asarray(pd.Series(y).dropna(),float)
    if not len(x) or not len(y):return np.nan
    u,_=mannwhitneyu(x,y,alternative="two-sided")
    return abs(1-2*u/(len(x)*len(y)))

def emotion_test(df,metric,group_col):
    valid=df[[group_col,metric]].dropna()
    groups={k:g[metric].dropna() for k,g in valid.groupby(group_col)}
    groups={k:v for k,v in groups.items() if len(v)>=2}
    base={"grouping":group_col,"metric":metric}
    if len(groups)<2:
        return {**base,"test_used":"not enough data","statistic":np.nan,"p_value":np.nan,
                "levene_p":np.nan,"normality_min_p":np.nan,"effect_size_name":np.nan,
                "effect_size":np.nan,"significant_0_05":np.nan}
    ps=[safe_shapiro(v) for v in groups.values()]
    ps=[p for p in ps if not np.isnan(p)]
    normal=min(ps) if ps else np.nan
    lev=safe_levene(list(groups.values()))
    anova=not np.isnan(normal) and not np.isnan(lev) and normal>ALPHA and lev>ALPHA
    try:
        if anova:
            stat,p=f_oneway(*groups.values()); test="ANOVA"
            ename="eta_squared"; effect=eta_squared(list(groups.values()))
        else:
            stat,p=kruskal(*groups.values()); test="Kruskal-Wallis"
            ename="epsilon_squared"; effect=epsilon_squared(stat,sum(map(len,groups.values())),len(groups))
    except:
        stat=p=effect=np.nan; test="error"; ename=np.nan
    return {**base,"test_used":test,"statistic":stat,"p_value":p,"levene_p":lev,
            "normality_min_p":normal,"effect_size_name":ename,"effect_size":effect,
            "significant_0_05":p<ALPHA if not np.isnan(p) else np.nan}

def dunn_bonferroni(df,metric,group_col):
    valid=df[[group_col,metric]].dropna()
    groups={k:g[metric].to_numpy(float) for k,g in valid.groupby(group_col) if len(g)>=2}
    labels=sorted(groups); pairs=list(combinations(labels,2))
    if not pairs:return []
    vals=np.concatenate([groups[k] for k in labels]); ranks=rankdata(vals)
    n=len(vals); _,tc=np.unique(vals,return_counts=True)
    tie=1-np.sum(tc**3-tc)/(n**3-n) if n>1 else 1
    var=n*(n+1)/12*tie
    means={}; off=0
    for k in labels:
        means[k]=ranks[off:off+len(groups[k])].mean(); off+=len(groups[k])
    out=[]
    for a,b in pairs:
        den=np.sqrt(var*(1/len(groups[a])+1/len(groups[b])))
        z=(means[a]-means[b])/den if den else np.nan
        raw=2*norm.sf(abs(z)) if not np.isnan(z) else np.nan
        adj=min(raw*len(pairs),1.) if not np.isnan(raw) else np.nan
        out.append({"grouping":group_col,"metric":metric,"omnibus_test":"Kruskal-Wallis",
                    "posthoc_test":"Dunn-Bonferroni","group_a":a,"group_b":b,
                    "statistic":z,"p_value_raw":raw,"p_value_adjusted":adj,
                    "pairwise_effect_size_name":"abs_rank_biserial",
                    "pairwise_effect_size":rank_biserial(groups[a],groups[b]),
                    "significant_0_05":adj<ALPHA if not np.isnan(adj) else np.nan})
    return out

def tukey(df,metric,group_col):
    v=df[[group_col,metric]].dropna()
    if v[group_col].nunique()<2:return []
    r=pairwise_tukeyhsd(v[metric].astype(float),v[group_col].astype(str),alpha=ALPHA)
    out=[]
    for x in r.summary().data[1:]:
        a,b,md,padj,lo,hi,reject=x
        out.append({"grouping":group_col,"metric":metric,"omnibus_test":"ANOVA",
                    "posthoc_test":"Tukey HSD","group_a":a,"group_b":b,"statistic":md,
                    "p_value_raw":np.nan,"p_value_adjusted":float(padj),
                    "pairwise_effect_size_name":np.nan,"pairwise_effect_size":np.nan,
                    "significant_0_05":bool(reject)})
    return out

def posthoc_table(df,omnibus):
    rows=[]
    for _,r in omnibus.iterrows():
        if r["significant_0_05"]!=True:continue
        if r["test_used"]=="ANOVA": rows+=tukey(df,r["metric"],r["grouping"])
        elif r["test_used"]=="Kruskal-Wallis": rows+=dunn_bonferroni(df,r["metric"],r["grouping"])
    return pd.DataFrame(rows)

def load_users():
    u=pd.read_excel(SUMMARY_FILE,header=None,usecols="A",skiprows=1,nrows=N_USERS)
    u.columns=["raw_name"]; u["participant"]=[f"U{i}" for i in range(1,len(u)+1)]
    u["clean_name"]=u["raw_name"].apply(clean_name); return u

def load_responses(users):
    rows=[]
    for _,u in users.iterrows():
        p=TODOS_DIR/f"respuestas_emociones_{u['clean_name']}.csv"
        if not p.exists():
            print("Missing:",p.name); continue
        d=pd.read_csv(p).rename(columns={"Respuesta":"response_value","Emocion":"response_emotion_raw","Archivo":"filename"})
        d["is_no_response"]=d["response_emotion_raw"].apply(is_no_response)
        d=d[~d["is_no_response"]].copy()
        d["response_emotion"]=d["response_emotion_raw"].apply(normalize_emotion)
        d=d.dropna(subset=["response_emotion"])
        d["response_value"]=pd.to_numeric(d["response_value"],errors="coerce")
        # Original recoding retained for comparability of the numeric scale:
        d.loc[d["response_value"]==4,"response_value"]=3
        d["participant"]=u["participant"]
        rows.append(d[["participant","filename","response_value","response_emotion"]])
    out=pd.concat(rows,ignore_index=True)
    if out.duplicated(["participant","filename"]).any():
        raise ValueError("Duplicate participant-image response rows found.")
    return out

def heatmap_inventory():
    rows=[]
    for p in sorted(HEATMAP_DIR.glob("*.npy")):
        try:
            a=np.load(p,mmap_mode="r"); h,w=a.shape if a.ndim==2 else (np.nan,np.nan)
        except:h=w=np.nan
        rows.append({"heatmap_name":p.stem,"heatmap_path":str(p),"heatmap_exists":True,"height":h,"width":w})
    return pd.DataFrame(rows)

def summarize(df,col):
    return df.dropna(subset=[col]).groupby(col,as_index=False).agg(
        observations=("filename","count"),images=("filename","nunique"),participants=("participant","nunique"),
        mean_fixation_count=("fixation_count","mean"),std_fixation_count=("fixation_count","std"),
        mean_dwell_time_s=("dwell_time_s","mean"),std_dwell_time_s=("dwell_time_s","std"),
        mean_fixation_duration_ms=("mean_fixation_duration_ms","mean"),
        std_fixation_duration_ms=("mean_fixation_duration_ms","std"),
        mean_fixation_density_per_megapixel=("fixation_density_per_megapixel","mean"),
        std_fixation_density_per_megapixel=("fixation_density_per_megapixel","std"))


def build_statistical_results_summary(user_df, omnibus_df, posthoc_df):
    """
    Creates a thesis-friendly summary with:
    omnibus result, significant pairwise comparisons, direction,
    group medians, adjusted p-value, and pairwise effect size.
    """
    rows = []

    for _, omnibus in omnibus_df.iterrows():
        grouping = omnibus["grouping"]
        metric = omnibus["metric"]

        valid = user_df[[grouping, metric]].dropna()

        group_medians = (
            valid.groupby(grouping)[metric]
            .median()
            .to_dict()
        )

        group_means = (
            valid.groupby(grouping)[metric]
            .mean()
            .to_dict()
        )

        relevant_posthoc = posthoc_df[
            (posthoc_df["grouping"] == grouping)
            & (posthoc_df["metric"] == metric)
        ].copy()

        if relevant_posthoc.empty:
            rows.append({
                "grouping": grouping,
                "metric": metric,
                "omnibus_test": omnibus["test_used"],
                "omnibus_statistic": omnibus["statistic"],
                "omnibus_p_value": omnibus["p_value"],
                "omnibus_effect_size_name": omnibus["effect_size_name"],
                "omnibus_effect_size": omnibus["effect_size"],
                "omnibus_significant_0_05": omnibus["significant_0_05"],
                "comparison": "No post-hoc test required",
                "group_a": np.nan,
                "group_b": np.nan,
                "group_a_median": np.nan,
                "group_b_median": np.nan,
                "group_a_mean": np.nan,
                "group_b_mean": np.nan,
                "direction_by_median": np.nan,
                "posthoc_test": np.nan,
                "adjusted_p_value": np.nan,
                "pairwise_effect_size_name": np.nan,
                "pairwise_effect_size": np.nan,
                "pairwise_significant_0_05": np.nan,
            })
            continue

        for _, pair in relevant_posthoc.iterrows():
            a = pair["group_a"]
            b = pair["group_b"]

            median_a = group_medians.get(a, np.nan)
            median_b = group_medians.get(b, np.nan)
            mean_a = group_means.get(a, np.nan)
            mean_b = group_means.get(b, np.nan)

            if pd.isna(median_a) or pd.isna(median_b):
                direction = np.nan
            elif median_a > median_b:
                direction = f"{a} > {b}"
            elif median_a < median_b:
                direction = f"{a} < {b}"
            else:
                direction = f"{a} = {b}"

            rows.append({
                "grouping": grouping,
                "metric": metric,
                "omnibus_test": omnibus["test_used"],
                "omnibus_statistic": omnibus["statistic"],
                "omnibus_p_value": omnibus["p_value"],
                "omnibus_effect_size_name": omnibus["effect_size_name"],
                "omnibus_effect_size": omnibus["effect_size"],
                "omnibus_significant_0_05": omnibus["significant_0_05"],
                "comparison": f"{a} vs {b}",
                "group_a": a,
                "group_b": b,
                "group_a_median": median_a,
                "group_b_median": median_b,
                "group_a_mean": mean_a,
                "group_b_mean": mean_b,
                "direction_by_median": direction,
                "posthoc_test": pair["posthoc_test"],
                "adjusted_p_value": pair["p_value_adjusted"],
                "pairwise_effect_size_name": pair["pairwise_effect_size_name"],
                "pairwise_effect_size": pair["pairwise_effect_size"],
                "pairwise_significant_0_05": pair["significant_0_05"],
            })

    return pd.DataFrame(rows)


def style_excel(path):
    wb=load_workbook(path); fill=PatternFill("solid",fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for c in ws[1]:c.font=Font(bold=True); c.fill=fill; c.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            m=max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(m+2,45)
    wb.save(path)

def main():
    aoi=pd.read_csv(AOI_FILE)
    if "emotion" in aoi.columns:aoi=aoi.rename(columns={"emotion":"original_emotion"})
    aoi["original_emotion"]=aoi["original_emotion"].apply(normalize_emotion)
    aoi["heatmap_name"]=aoi["trial_index"].apply(lambda x:f"heatmap_{int(x):02d}")
    aoi["image_area_px"]=pd.to_numeric(aoi["orig_w"],errors="coerce")*pd.to_numeric(aoi["orig_h"],errors="coerce")
    aoi["image_area_megapixels"]=aoi["image_area_px"]/1_000_000

    usable=pd.read_excel(QA_FILE,sheet_name=QA_SHEET,usecols="B:AM",nrows=N_IMAGES)
    responses=load_responses(load_users())
    rows=[]
    for participant in usable.columns:
        fp=RESULTS_DIR/participant/"valid_fixations.csv"
        if not fp.exists():continue
        fix=pd.read_csv(fp)
        for _,img in aoi.iterrows():
            trial=int(img["trial_index"])
            if usable.loc[trial-1,participant]!=1:continue
            f=fix[fix["filename"]==img["filename"]]
            count=len(f); dwell=f["duration"].sum() if count else 0.
            mean_dur=f["duration"].mean() if count else np.nan
            area=img["image_area_megapixels"]
            density=count/area if pd.notna(area) and area>0 else np.nan
            rows.append({"participant":participant,"trial_index":trial,"filename":img["filename"],
                         "heatmap_name":img["heatmap_name"],"original_emotion":img["original_emotion"],
                         "fixation_count":count,"dwell_time_s":dwell/1000,
                         "mean_fixation_duration_ms":mean_dur,
                         "fixation_density_per_megapixel":density})
    user=pd.DataFrame(rows).merge(responses,on=["participant","filename"],how="inner",validate="one_to_one")
    user["emotion_match_original_vs_response"]=user["original_emotion"]==user["response_emotion"]

    image=user.groupby(["trial_index","filename","heatmap_name","original_emotion"],as_index=False).agg(
        usable_participants=("participant","nunique"),total_fixations=("fixation_count","sum"),
        mean_fixations_per_user=("fixation_count","mean"),std_fixations_per_user=("fixation_count","std"),
        total_dwell_time_s=("dwell_time_s","sum"),mean_dwell_time_s_per_user=("dwell_time_s","mean"),
        std_dwell_time_s_per_user=("dwell_time_s","std"),
        mean_fixation_duration_ms=("mean_fixation_duration_ms","mean"),
        mean_fixation_density_per_megapixel=("fixation_density_per_megapixel","mean"),
        most_common_response_emotion=("response_emotion",lambda x:x.mode().iloc[0] if not x.mode().empty else np.nan),
        response_match_rate=("emotion_match_original_vs_response","mean"))

    inv=heatmap_inventory()
    if not inv.empty:
        image=image.merge(inv,on="heatmap_name",how="left",validate="one_to_one")

    orig_sum=summarize(user,"original_emotion")
    resp_sum=summarize(user,"response_emotion")
    combined=user.dropna(subset=["original_emotion","response_emotion"]).groupby(
        ["original_emotion","response_emotion"],as_index=False).agg(
        observations=("filename","count"),images=("filename","nunique"),participants=("participant","nunique"),
        mean_fixation_count=("fixation_count","mean"),mean_dwell_time_s=("dwell_time_s","mean"),
        mean_fixation_duration_ms=("mean_fixation_duration_ms","mean"),
        mean_fixation_density_per_megapixel=("fixation_density_per_megapixel","mean"))
    confusion=pd.crosstab(user["original_emotion"],user["response_emotion"],margins=True).reset_index()

    metrics=["fixation_count","dwell_time_s","mean_fixation_duration_ms","fixation_density_per_megapixel"]
    stat_rows=[]
    for m in metrics:
        stat_rows.append(emotion_test(user,m,"original_emotion"))
        stat_rows.append(emotion_test(user,m,"response_emotion"))
    stats=pd.DataFrame(stat_rows)
    posthoc=posthoc_table(user,stats)
    statistical_summary=build_statistical_results_summary(user,stats,posthoc)

    participant=user.groupby("participant",as_index=False).agg(
        usable_images=("filename","nunique"),total_fixations=("fixation_count","sum"),
        total_dwell_time_s=("dwell_time_s","sum"),mean_fixations_per_image=("fixation_count","mean"),
        mean_fixation_duration_ms=("mean_fixation_duration_ms","mean"),
        response_match_rate=("emotion_match_original_vs_response","mean"))

    summary=pd.DataFrame({"metric":["Participants included","Images included","Usable participant-image observations",
        "Total fixations","Total dwell time seconds","Mean fixation duration ms","Response files loaded",
        "Mean original-response match rate","NO RESPONDE handling","Response value 4 handling","Heatmap folder"],
        "value":[user["participant"].nunique(),user["filename"].nunique(),len(user),user["fixation_count"].sum(),
        user["dwell_time_s"].sum(),user["mean_fixation_duration_ms"].mean(),responses["participant"].nunique(),
        user["emotion_match_original_vs_response"].mean(),"Excluded before analysis","Recoded as 3",str(HEATMAP_DIR)]})

    with pd.ExcelWriter(OUTPUT_FILE,engine="openpyxl") as w:
        summary.to_excel(w,sheet_name="Summary",index=False)
        image.to_excel(w,sheet_name="Image Metrics",index=False)
        user.to_excel(w,sheet_name="User Image Metrics",index=False)
        participant.to_excel(w,sheet_name="Participant Summary",index=False)
        orig_sum.to_excel(w,sheet_name="Original Emotion Summary",index=False)
        resp_sum.to_excel(w,sheet_name="Response Emotion Summary",index=False)
        combined.to_excel(w,sheet_name="Combined Emotion Summary",index=False)
        confusion.to_excel(w,sheet_name="Emotion Confusion",index=False)
        stats.to_excel(w,sheet_name="Emotion Statistics",index=False)
        posthoc.to_excel(w,sheet_name="Posthoc Tests",index=False)
        statistical_summary.to_excel(w,sheet_name="Statistical Results Summary",index=False)
        if not inv.empty:inv.to_excel(w,sheet_name="Heatmap Inventory",index=False)
    style_excel(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")

if __name__=="__main__":
    main()
