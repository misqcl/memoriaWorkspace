from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
from scipy.ndimage import zoom
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent

PRED_DIR = BASE_DIR / "deepgaze_predictions" / "raw"
ANALYSIS_DIR = BASE_DIR / "analisis_2"
VALID_PAIRS_FILE = ANALYSIS_DIR / "valid_response_observations.csv"
EMPIRICAL_HEATMAP_DIR = ANALYSIS_DIR / "image_heatmaps" / "raw"
RESULTS_DIR = BASE_DIR / "results"
IMAGE_AOI_FILE = BASE_DIR / "image_aois" / "image_aoi_seed42.csv"
QA_FILE = ANALYSIS_DIR / "valid_response_matrix.xlsx"
QA_SHEET = "Usable"

OUTPUT_DIR = ANALYSIS_DIR / "deepgaze_evaluation"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

BY_IMAGE_CSV = OUTPUT_DIR / "deepgaze_metrics_by_image.csv"
SUMMARY_CSV = OUTPUT_DIR / "deepgaze_metrics_summary.csv"
XLSX_FILE = OUTPUT_DIR / "deepgaze_evaluation.xlsx"

EPS = 1e-12


def normalize_prob(x):
    x = np.asarray(x, dtype=np.float64)
    x[~np.isfinite(x)] = 0
    x = np.maximum(x, 0)
    s = x.sum()
    return x / s if s > 0 else np.zeros_like(x)


def resize_to_shape(arr, shape):
    if arr.shape == shape:
        return arr.astype(np.float64)
    factors = (shape[0] / arr.shape[0], shape[1] / arr.shape[1])
    return zoom(arr, factors, order=1)


def cc(pred, empirical):
    p = np.asarray(pred, float).ravel()
    e = np.asarray(empirical, float).ravel()
    if np.std(p) == 0 or np.std(e) == 0:
        return np.nan
    return float(np.corrcoef(p, e)[0, 1])


def sim(pred, empirical):
    p = normalize_prob(pred)
    e = normalize_prob(empirical)
    return float(np.minimum(p, e).sum())


def kl_divergence(empirical, pred):
    # KL(empirical || prediction): information lost when prediction approximates empirical data.
    p = normalize_prob(pred)
    q = normalize_prob(empirical)
    mask = q > 0
    return float(np.sum(q[mask] * np.log((q[mask] + EPS) / (p[mask] + EPS))))


def nss(pred, fixation_map):
    s = np.asarray(pred, float)
    sd = s.std()
    if sd == 0:
        return np.nan
    z = (s - s.mean()) / sd
    ys, xs = np.nonzero(fixation_map > 0)
    if len(xs) == 0:
        return np.nan
    weights = fixation_map[ys, xs].astype(float)
    return float(np.average(z[ys, xs], weights=weights))


def auc_judd(pred, fixation_map):
    sal = np.asarray(pred, float)
    fix = np.asarray(fixation_map) > 0

    if fix.sum() == 0 or (~fix).sum() == 0:
        return np.nan

    sal = sal.copy()
    mn, mx = sal.min(), sal.max()
    if mx > mn:
        sal = (sal - mn) / (mx - mn)

    fix_vals = sal[fix]
    thresholds = np.sort(np.unique(fix_vals))[::-1]

    tp = [0.0]
    fp = [0.0]
    n_fix = fix.sum()
    n_nonfix = (~fix).sum()

    for thresh in thresholds:
        above = sal >= thresh
        tp.append(float((above & fix).sum()) / n_fix)
        fp.append(float((above & ~fix).sum()) / n_nonfix)

    tp.append(1.0)
    fp.append(1.0)

    order = np.argsort(fp)
    return float(np.trapezoid(np.asarray(tp)[order], np.asarray(fp)[order]))


def find_prediction(trial, filename):
    stem = Path(str(filename)).stem
    exact = PRED_DIR / f"{int(trial):02d}_{stem}_deepgaze_density.npy"
    if exact.exists():
        return exact
    matches = list(PRED_DIR.glob(f"*_{stem}_deepgaze_density.npy"))
    return matches[0] if matches else None


def find_empirical_heatmap(filename, trial):
    stem = Path(str(filename)).stem
    candidates = [
        EMPIRICAL_HEATMAP_DIR / f"{stem}.npy",
        EMPIRICAL_HEATMAP_DIR / f"heatmap_{int(trial):02d}.npy",
    ]
    for p in candidates:
        if p.exists():
            return p
    matches = list(EMPIRICAL_HEATMAP_DIR.glob(f"*{stem}*.npy"))
    return matches[0] if matches else None


def load_inputs():
    image_info = pd.read_csv(IMAGE_AOI_FILE).sort_values("trial_index")
    if "emotion" in image_info.columns and "original_emotion" not in image_info.columns:
        image_info = image_info.rename(columns={"emotion": "original_emotion"})

    usable_raw = pd.read_excel(
        QA_FILE,
        sheet_name=QA_SHEET,
        nrows=len(image_info),
    )
    usable_raw = usable_raw.dropna(axis=1, how="all")
    participant_columns = [
        column
        for column in usable_raw.columns
        if str(column).strip().upper().startswith("U")
    ]
    if not participant_columns:
        raise ValueError(
            f"No participant columns were found in sheet '{QA_SHEET}'. "
            f"Available columns: {list(usable_raw.columns)}"
        )

    usable = usable_raw[participant_columns].copy()
    return image_info, usable


def build_fixation_map(row, usable, target_shape):
    trial = int(row["trial_index"])
    filename = str(row["filename"])
    h, w = target_shape

    left = float(row["surf_left"])
    right = float(row["surf_right"])
    top = float(row["surf_top"])
    bottom = float(row["surf_bottom"])

    fixation_map = np.zeros((h, w), dtype=np.float64)

    for participant in usable.columns:
        try:
            if usable.loc[trial - 1, participant] != 1:
                continue
        except Exception:
            continue

        csv_path = RESULTS_DIR / str(participant) / "valid_fixations.csv"
        if not csv_path.exists():
            continue

        fix = pd.read_csv(csv_path, usecols=lambda c: c in {
            "filename", "norm_pos_x", "norm_pos_y"
        })
        if fix.empty or "filename" not in fix.columns:
            continue

        fix = fix[fix["filename"].astype(str) == filename]
        for _, f in fix.iterrows():
            sx = pd.to_numeric(f["norm_pos_x"], errors="coerce")
            sy = pd.to_numeric(f["norm_pos_y"], errors="coerce")
            if pd.isna(sx) or pd.isna(sy):
                continue

            ix = (float(sx) - left) / (right - left)
            iy = (float(sy) - top) / (bottom - top)

            if 0 <= ix <= 1 and 0 <= iy <= 1:
                x = min(w - 1, max(0, int(round(ix * (w - 1)))))
                y = min(h - 1, max(0, int(round(iy * (h - 1)))))
                fixation_map[y, x] += 1

    return fixation_map


def bootstrap_ci(values, n_boot=5000, seed=42):
    vals = np.asarray(pd.Series(values).dropna(), float)
    if len(vals) < 2:
        return np.nan, np.nan
    rng = np.random.default_rng(seed)
    means = np.empty(n_boot)
    for i in range(n_boot):
        means[i] = rng.choice(vals, size=len(vals), replace=True).mean()
    return float(np.percentile(means, 2.5)), float(np.percentile(means, 97.5))


def style_excel(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 42)
    wb.save(path)


def main():
    image_info, usable = load_inputs()
    rows = []

    print(f"Evaluating {len(image_info)} images...")

    for i, (_, row) in enumerate(image_info.iterrows(), start=1):
        trial = int(row["trial_index"])
        filename = str(row["filename"])

        pred_path = find_prediction(trial, filename)
        emp_path = find_empirical_heatmap(filename, trial)

        if pred_path is None or emp_path is None:
            print(f"{i:02d}/{len(image_info)} {filename}: missing prediction or empirical heatmap")
            continue

        pred = np.load(pred_path)
        empirical = np.load(emp_path)

        pred = resize_to_shape(pred, empirical.shape)
        pred = normalize_prob(pred)
        empirical = normalize_prob(empirical)

        fix_map = build_fixation_map(row, usable, empirical.shape)

        result = {
            "trial_index": trial,
            "filename": filename,
            "original_emotion": row.get("original_emotion", np.nan),
            "n_fixations_for_auc_nss": int(fix_map.sum()),
            "CC": cc(pred, empirical),
            "SIM": sim(pred, empirical),
            "KL_empirical_to_prediction": kl_divergence(empirical, pred),
            "AUC_Judd": auc_judd(pred, fix_map),
            "NSS": nss(pred, fix_map),
            "prediction_path": str(pred_path),
            "empirical_heatmap_path": str(emp_path),
        }
        rows.append(result)

        print(
            f"{i:02d}/{len(image_info)} {filename}: "
            f"CC={result['CC']:.3f}, SIM={result['SIM']:.3f}, "
            f"KL={result['KL_empirical_to_prediction']:.3f}, "
            f"AUC={result['AUC_Judd']:.3f}, NSS={result['NSS']:.3f}"
        )

    metrics = pd.DataFrame(rows)
    if metrics.empty:
        raise RuntimeError(
            "No images were evaluated. Check deepgaze_predictions/raw and image_heatmaps/raw."
        )

    metric_names = ["CC", "SIM", "KL_empirical_to_prediction", "AUC_Judd", "NSS"]
    summary_rows = []
    for metric in metric_names:
        lo, hi = bootstrap_ci(metrics[metric])
        summary_rows.append({
            "metric": metric,
            "n_images": int(metrics[metric].notna().sum()),
            "mean": metrics[metric].mean(),
            "std": metrics[metric].std(),
            "median": metrics[metric].median(),
            "min": metrics[metric].min(),
            "max": metrics[metric].max(),
            "bootstrap_95CI_mean_low": lo,
            "bootstrap_95CI_mean_high": hi,
        })
    summary = pd.DataFrame(summary_rows)

    emotion_summary = (
        metrics.groupby("original_emotion", dropna=False)[metric_names]
        .agg(["count", "mean", "std", "median"])
        .reset_index()
    )
    emotion_summary.columns = [
        "_".join([str(x) for x in col if str(x)])
        if isinstance(col, tuple) else str(col)
        for col in emotion_summary.columns
    ]

    metrics.to_csv(BY_IMAGE_CSV, index=False, encoding="utf-8-sig")
    summary.to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="Metrics by Image", index=False)
        summary.to_excel(writer, sheet_name="Overall Summary", index=False)
        emotion_summary.to_excel(writer, sheet_name="By Original Emotion", index=False)

    style_excel(XLSX_FILE)

    print("\nDone.")
    print(f"By-image metrics: {BY_IMAGE_CSV}")
    print(f"Summary:          {SUMMARY_CSV}")
    print(f"Workbook:         {XLSX_FILE}")


if __name__ == "__main__":
    main()
